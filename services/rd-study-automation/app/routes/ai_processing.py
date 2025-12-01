"""
AI Data Processing Routes

Handles AI-powered document processing, data extraction, and study completion.
Uses OpenAI for intelligent Excel parsing and data extraction.
"""

import logging
import io
import json
import os
from uuid import UUID
from datetime import datetime
from typing import List, Optional
from decimal import Decimal

from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form, BackgroundTasks, Request
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, update
import openpyxl
import pandas as pd

from ..database import get_db
from ..config import settings
from ..models import (
    RDStudy, RDProject, RDEmployee, QualifiedResearchExpense, RDDocument,
    RDCalculation, StudyStatus, QualificationStatus, QRECategory, CreditMethod
)

# OpenAI client for AI-powered parsing
try:
    from openai import OpenAI
    openai_client = OpenAI(api_key=os.getenv("OPENAI_API_KEY", settings.OPENAI_API_KEY if hasattr(settings, 'OPENAI_API_KEY') else ""))
    HAS_OPENAI = bool(os.getenv("OPENAI_API_KEY") or (hasattr(settings, 'OPENAI_API_KEY') and settings.OPENAI_API_KEY))
except Exception:
    openai_client = None
    HAS_OPENAI = False

logger = logging.getLogger(__name__)

router = APIRouter()


# =============================================================================
# UNIFIED FILE UPLOAD & ANALYSIS (Expected by Frontend)
# =============================================================================

async def ai_analyze_columns(columns: list, sample_data: list) -> dict:
    """Use OpenAI to intelligently analyze columns and detect data types."""
    if not HAS_OPENAI or not openai_client:
        return None

    try:
        # Prepare data summary for AI
        data_summary = f"Columns: {columns}\n\nSample data (first 3 rows):\n"
        for i, row in enumerate(sample_data[:3]):
            data_summary += f"Row {i+1}: {json.dumps(row, default=str)}\n"

        prompt = f"""Analyze this Excel/CSV data for R&D tax credit study purposes.

{data_summary}

Determine:
1. What type of data this is: "payroll" (employee wages), "projects" (R&D projects), or "expenses" (supplies/contracts)
2. Map each column to one of these fields:
   - For payroll: name, title, department, wages, qualified_time
   - For projects: name, description, department, business_component, start_date, end_date
   - For expenses: description, amount, vendor, category

Return JSON in this exact format:
{{
    "data_type": "payroll|projects|expenses",
    "confidence": 0.0-1.0,
    "column_mappings": [
        {{"source_column": "column_name", "suggested_field": "field_name", "confidence": 0.0-1.0}},
        ...
    ],
    "analysis": "Brief explanation of what you found",
    "recommendations": ["recommendation 1", "recommendation 2"]
}}"""

        response = openai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.3,
            max_tokens=1000,
            response_format={"type": "json_object"}
        )

        result = json.loads(response.choices[0].message.content)
        return result
    except Exception as e:
        logger.warning(f"OpenAI analysis failed: {str(e)}")
        return None


def fallback_column_analysis(columns: list, sample_data: list) -> dict:
    """Fallback pattern-based column analysis when AI is unavailable."""
    column_mappings = []
    detected_types = []

    # Pattern indicators
    payroll_indicators = {
        'name': ['name', 'employee', 'full_name', 'employee_name', 'emp_name', 'worker', 'staff'],
        'title': ['title', 'job_title', 'position', 'role', 'job', 'designation'],
        'department': ['department', 'dept', 'division', 'cost_center', 'team', 'group', 'unit'],
        'wages': ['wages', 'salary', 'w2', 'gross_pay', 'total_wages', 'compensation', 'earnings', 'pay', 'income', 'annual', 'ytd'],
        'qualified_time': ['qualified', 'rd_time', 'r&d', 'research_pct', 'percent', 'qre', '%']
    }

    project_indicators = {
        'name': ['project', 'project_name', 'initiative', 'program'],
        'description': ['description', 'desc', 'details', 'summary', 'overview', 'notes', 'activities'],
        'business_component': ['business_component', 'component', 'product', 'service', 'business']
    }

    expense_indicators = {
        'description': ['description', 'desc', 'item', 'expense', 'memo', 'purpose'],
        'amount': ['amount', 'total', 'cost', 'value', 'price', 'sum', 'spend'],
        'vendor': ['vendor', 'supplier', 'payee', 'company', 'contractor']
    }

    for col in columns:
        col_lower = str(col).lower().strip().replace('_', ' ')
        mapping = {"source_column": str(col), "confidence": 0.0, "suggested_field": None, "data_type": None}

        # Check payroll indicators
        for field, indicators in payroll_indicators.items():
            for ind in indicators:
                if ind in col_lower or col_lower in ind:
                    mapping["suggested_field"] = field
                    mapping["confidence"] = 0.9 if ind == col_lower or col_lower == ind else 0.75
                    mapping["data_type"] = "payroll"
                    if "payroll" not in detected_types:
                        detected_types.append("payroll")
                    break
            if mapping["suggested_field"]:
                break

        # Check project indicators
        if not mapping["suggested_field"]:
            for field, indicators in project_indicators.items():
                for ind in indicators:
                    if ind in col_lower or col_lower in ind:
                        mapping["suggested_field"] = field
                        mapping["confidence"] = 0.85 if ind == col_lower else 0.7
                        mapping["data_type"] = "projects"
                        if "projects" not in detected_types:
                            detected_types.append("projects")
                        break
                if mapping["suggested_field"]:
                    break

        # Check expense indicators
        if not mapping["suggested_field"]:
            for field, indicators in expense_indicators.items():
                for ind in indicators:
                    if ind in col_lower or col_lower in ind:
                        mapping["suggested_field"] = field
                        mapping["confidence"] = 0.80 if ind == col_lower else 0.65
                        mapping["data_type"] = "expenses"
                        if "supplies" not in detected_types:
                            detected_types.append("supplies")
                        break
                if mapping["suggested_field"]:
                    break

        column_mappings.append(mapping)

    # Infer from sample data if column names don't match
    for i, col in enumerate(columns):
        if column_mappings[i]["suggested_field"] is None and sample_data:
            sample_values = [str(row.get(col, '')) for row in sample_data[:5] if row.get(col)]
            # Check if values look like currency/wages
            currency_count = sum(1 for v in sample_values if '$' in v or any(c.isdigit() for c in v.replace(',', '').replace('.', '')))
            if currency_count >= len(sample_values) * 0.7 and sample_values:
                # Check if large numbers (likely wages)
                try:
                    numeric_vals = [float(v.replace('$', '').replace(',', '')) for v in sample_values if v]
                    if any(n > 1000 for n in numeric_vals):
                        column_mappings[i]["suggested_field"] = "wages"
                        column_mappings[i]["confidence"] = 0.6
                        column_mappings[i]["data_type"] = "payroll"
                        if "payroll" not in detected_types:
                            detected_types.append("payroll")
                except:
                    pass

    primary_type = "payroll" if "payroll" in detected_types else ("projects" if "projects" in detected_types else ("expenses" if detected_types else "unknown"))

    return {
        "data_type": primary_type,
        "column_mappings": column_mappings,
        "detected_types": detected_types
    }


@router.post("/studies/{study_id}/upload/analyze")
async def analyze_uploaded_file(
    study_id: UUID,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db)
):
    """
    AI-powered analysis of uploaded Excel/CSV file.
    Uses OpenAI GPT-4 to intelligently detect data types, column mappings, and extract information.
    Falls back to pattern matching if AI is unavailable.
    """
    study = await db.get(RDStudy, study_id)
    if not study:
        raise HTTPException(status_code=404, detail="Study not found")

    # Read file content
    content = await file.read()
    filename = file.filename or "unknown"
    sheets_info = []

    try:
        # Detect file type and parse
        if filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(content))
            sheets_info = [{"name": "Sheet1", "row_count": len(df), "columns": list(df.columns)}]
        else:
            wb = openpyxl.load_workbook(io.BytesIO(content))
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                row_count = sum(1 for _ in sheet.iter_rows(min_row=2, values_only=True) if any(cell for cell in _))
                cols = [cell.value for cell in sheet[1] if cell.value]
                sheets_info.append({
                    "name": sheet_name,
                    "row_count": row_count,
                    "columns": cols
                })
            # Use first sheet for analysis
            df = pd.read_excel(io.BytesIO(content), sheet_name=wb.sheetnames[0])

        # Clean column names
        df.columns = [str(c).strip() for c in df.columns]
        columns = list(df.columns)

        # Get all data (not just sample) for import
        all_data = df.replace({pd.NaT: None, float('nan'): None}).to_dict('records')
        sample_data = all_data[:5]

        # Try AI-powered analysis first
        ai_result = await ai_analyze_columns(columns, sample_data)

        if ai_result:
            # Use AI analysis
            logger.info("Using AI-powered column analysis")
            column_mappings = ai_result.get("column_mappings", [])
            primary_type = ai_result.get("data_type", "unknown")
            detected_types = [primary_type] if primary_type != "unknown" else []
            ai_analysis = ai_result.get("analysis", "")
            ai_recommendations = ai_result.get("recommendations", [])
            overall_confidence = ai_result.get("confidence", 0.8)
            used_ai = True
        else:
            # Fallback to pattern matching
            logger.info("Using fallback pattern-based column analysis")
            fallback_result = fallback_column_analysis(columns, sample_data)
            column_mappings = fallback_result["column_mappings"]
            primary_type = fallback_result["data_type"]
            detected_types = fallback_result["detected_types"]
            ai_analysis = ""
            ai_recommendations = []
            mapped_count = sum(1 for m in column_mappings if m["confidence"] > 0.5)
            overall_confidence = mapped_count / len(columns) if columns else 0
            used_ai = False

        # Generate issues
        issues = []
        if overall_confidence < 0.5:
            issues.append({
                "type": "low_confidence",
                "message": "Low column mapping confidence. Please verify mappings before import.",
                "severity": "warning"
            })

        # Check for missing critical fields
        if primary_type == "payroll":
            found = [m.get("suggested_field") for m in column_mappings if m.get("confidence", 0) > 0.5]
            missing = [r for r in ["name", "wages"] if r not in found]
            if missing:
                issues.append({
                    "type": "missing_fields",
                    "message": f"Missing recommended fields: {', '.join(missing)}",
                    "severity": "warning"
                })

        # Build recommendations
        recommendations = ai_recommendations if ai_recommendations else [
            f"Detected as {primary_type} data with {overall_confidence:.0%} confidence",
            "Review column mappings before importing",
            "Ensure all required fields are mapped correctly"
        ]

        return {
            "success": True,
            "filename": filename,
            "file_size": len(content),
            "sheets": sheets_info,
            "detected_data_types": detected_types,
            "primary_data_type": primary_type,
            "columns": columns,
            "column_mappings": column_mappings,
            "sample_data": sample_data,
            "all_data": all_data,  # Include all data for import
            "total_rows": len(df),
            "overall_confidence": overall_confidence,
            "issues": issues,
            "recommendations": recommendations,
            "ai_analysis": ai_analysis,
            "used_ai": used_ai
        }

    except Exception as e:
        logger.error(f"Error analyzing file: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Error analyzing file: {str(e)}")


@router.post("/studies/{study_id}/upload/import")
async def import_analyzed_data(
    study_id: UUID,
    import_request: dict,
    db: AsyncSession = Depends(get_db)
):
    """
    Import data based on analyzed column mappings.
    """
    study = await db.get(RDStudy, study_id)
    if not study:
        raise HTTPException(status_code=404, detail="Study not found")

    data_type = import_request.get("data_type", "payroll")
    mappings = import_request.get("mappings", {})
    data = import_request.get("data", [])

    imported_count = 0
    errors = []

    try:
        if data_type == "payroll":
            # Import as employees
            for row in data:
                try:
                    name = row.get(mappings.get("name", "name"), "")
                    if not name:
                        continue

                    wages_str = str(row.get(mappings.get("wages", "wages"), "0"))
                    wages_str = wages_str.replace("$", "").replace(",", "").strip()
                    w2_wages = Decimal(wages_str) if wages_str else Decimal("0")

                    qualified_pct_str = str(row.get(mappings.get("qualified_time", ""), "0"))
                    qualified_pct_str = qualified_pct_str.replace("%", "").strip()
                    try:
                        qualified_pct = Decimal(qualified_pct_str) if qualified_pct_str else Decimal("50")
                        if qualified_pct > 100:
                            qualified_pct = Decimal("50")
                    except:
                        qualified_pct = Decimal("50")

                    employee = RDEmployee(
                        study_id=study_id,
                        name=str(name).strip(),
                        title=str(row.get(mappings.get("title", ""), "")).strip() or None,
                        department=str(row.get(mappings.get("department", ""), "")).strip() or None,
                        w2_wages=w2_wages,
                        total_wages=w2_wages,
                        qualified_time_percentage=qualified_pct,
                        qualified_wages=w2_wages * (qualified_pct / 100),
                        qualified_time_source="excel_import"
                    )
                    db.add(employee)
                    imported_count += 1
                except Exception as e:
                    errors.append(f"Row error: {str(e)}")

        elif data_type == "projects":
            for row in data:
                try:
                    name = row.get(mappings.get("name", "name"), "")
                    if not name:
                        continue

                    project = RDProject(
                        study_id=study_id,
                        name=str(name).strip(),
                        description=str(row.get(mappings.get("description", ""), "")).strip() or None,
                        business_component=str(row.get(mappings.get("business_component", ""), "")).strip() or None,
                        qualification_status=QualificationStatus.PENDING
                    )
                    db.add(project)
                    imported_count += 1
                except Exception as e:
                    errors.append(f"Row error: {str(e)}")

        elif data_type in ["supplies", "expenses"]:
            for row in data:
                try:
                    desc = row.get(mappings.get("description", "description"), "")
                    amount_str = str(row.get(mappings.get("amount", "amount"), "0"))
                    amount_str = amount_str.replace("$", "").replace(",", "").strip()
                    try:
                        gross_amount = Decimal(amount_str)
                    except:
                        continue

                    qre = QualifiedResearchExpense(
                        study_id=study_id,
                        category=QRECategory.SUPPLIES,
                        description=str(desc).strip() or "Imported supply",
                        supply_vendor=str(row.get(mappings.get("vendor", ""), "")).strip() or None,
                        gross_amount=gross_amount,
                        qualified_percentage=Decimal("100"),
                        qualified_amount=gross_amount
                    )
                    db.add(qre)
                    imported_count += 1
                except Exception as e:
                    errors.append(f"Row error: {str(e)}")

        await db.commit()

        return {
            "success": True,
            "imported_count": imported_count,
            "data_type": data_type,
            "errors": errors[:10] if errors else [],
            "message": f"Successfully imported {imported_count} {data_type} records"
        }

    except Exception as e:
        logger.error(f"Error importing data: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Error importing data: {str(e)}")


@router.post("/studies/{study_id}/payroll/connect")
async def connect_payroll_provider(
    study_id: UUID,
    connection: dict,
    db: AsyncSession = Depends(get_db)
):
    """Connect to a payroll provider for automated data import."""
    study = await db.get(RDStudy, study_id)
    if not study:
        raise HTTPException(status_code=404, detail="Study not found")

    provider = connection.get("provider", "").lower()
    # Map frontend provider IDs to canonical names
    provider_mapping = {
        "adp_run": "adp",
        "adp": "adp",
        "justworks": "justworks",
        "paychex": "paychex",
        "paychex_flex": "paychex",
        "gusto": "gusto",
        "quickbooks": "quickbooks",
        "sage": "sage",
        "intuit": "intuit"
    }

    canonical_provider = provider_mapping.get(provider)
    if not canonical_provider:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported provider '{provider}'. Supported: adp_run, justworks, paychex"
        )
    provider = canonical_provider

    # Store connection config (would be encrypted in production)
    if not study.ai_suggested_areas:
        study.ai_suggested_areas = {}
    study.ai_suggested_areas["payroll_connection"] = {
        "provider": provider,
        "status": "pending_oauth",
        "configured_at": datetime.utcnow().isoformat()
    }

    await db.commit()

    # Return OAuth URL (in production, would generate actual OAuth URL)
    return {
        "success": True,
        "provider": provider,
        "status": "pending_oauth",
        "oauth_url": f"/api/rd-study/oauth/{provider}/authorize?study_id={study_id}",
        "message": f"Please complete OAuth authentication with {provider}"
    }


# =============================================================================
# AI HELPER FUNCTIONS
# =============================================================================

async def qualify_project_with_ai(openai_client, project: RDProject) -> dict:
    """
    Use OpenAI to analyze a project against the IRS 4-part test for R&D qualification.
    Returns scores and analysis for each part of the test.
    """
    if not openai_client:
        # Fallback for when OpenAI is not configured
        return _fallback_qualify_project(project)

    # Determine model/deployment name based on client type
    if settings.AZURE_OPENAI_ENDPOINT and settings.AZURE_OPENAI_API_KEY:
        model_name = settings.AZURE_OPENAI_DEPLOYMENT or "gpt-4-turbo"
    else:
        model_name = settings.OPENAI_CHAT_MODEL

    prompt = f"""Analyze this R&D project against the IRS Section 41 four-part test for R&D tax credit qualification.

PROJECT INFORMATION:
- Name: {project.name}
- Description: {project.description or 'No description provided'}
- Business Component: {project.business_component or 'Not specified'}
- Department: {project.department or 'Not specified'}

Evaluate each of the four parts of the test on a scale of 0-100:

1. PERMITTED PURPOSE: Does the activity intend to create or improve a product, process, software, technique, formula, or invention?

2. TECHNOLOGICAL IN NATURE: Is the activity fundamentally technological, relying on principles of physical science, biological science, engineering, or computer science?

3. ELIMINATION OF UNCERTAINTY: Does the activity seek to eliminate uncertainty about capability, method, or design of the business component?

4. PROCESS OF EXPERIMENTATION: Does the activity involve evaluating alternatives through modeling, simulation, systematic trial and error, or other methods?

Respond with valid JSON only (no markdown):
{{
    "permitted_purpose_score": <0-100>,
    "permitted_purpose_analysis": "<explanation>",
    "technological_nature_score": <0-100>,
    "technological_nature_analysis": "<explanation>",
    "uncertainty_score": <0-100>,
    "uncertainty_analysis": "<explanation>",
    "experimentation_score": <0-100>,
    "experimentation_analysis": "<explanation>",
    "overall_score": <0-100>,
    "qualification_status": "<qualified|needs_review|not_qualified>",
    "qualification_narrative": "<summary of qualification>",
    "risk_flags": ["<list of any audit concerns>"],
    "suggested_evidence": ["<list of evidence that would strengthen the case>"]
}}"""

    try:
        response = await openai_client.chat.completions.create(
            model=model_name,
            messages=[
                {"role": "system", "content": "You are an expert R&D tax credit analyst with deep knowledge of IRC Section 41, Treasury Regulations, and IRS guidance. Provide objective, audit-defensible analysis."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.2,
            max_tokens=1500
        )

        result_text = response.choices[0].message.content.strip()
        # Clean up potential markdown formatting
        if result_text.startswith("```"):
            result_text = result_text.split("```")[1]
            if result_text.startswith("json"):
                result_text = result_text[4:]
        result = json.loads(result_text)
        logger.info(f"AI project qualification successful using {model_name}")
        return result

    except Exception as e:
        logger.error(f"OpenAI project qualification failed: {e}")
        return _fallback_qualify_project(project)


def _fallback_qualify_project(project: RDProject) -> dict:
    """Fallback qualification using rule-based analysis when OpenAI is unavailable."""
    desc = (project.description or "").lower()
    base_score = 70

    # Adjust based on keywords
    if any(kw in desc for kw in ['develop', 'create', 'design', 'engineer', 'build']):
        base_score += 10
    if any(kw in desc for kw in ['new', 'novel', 'innovative', 'breakthrough']):
        base_score += 5
    if any(kw in desc for kw in ['test', 'experiment', 'prototype', 'iterate']):
        base_score += 5
    if any(kw in desc for kw in ['uncertain', 'challenge', 'solve', 'problem']):
        base_score += 5

    if base_score >= 75:
        status = "qualified"
    elif base_score >= 50:
        status = "needs_review"
    else:
        status = "not_qualified"

    return {
        "permitted_purpose_score": min(100, base_score + 5),
        "permitted_purpose_analysis": "Analysis requires AI configuration. Based on keywords detected.",
        "technological_nature_score": min(100, base_score),
        "technological_nature_analysis": "Analysis requires AI configuration. Based on keywords detected.",
        "uncertainty_score": min(100, base_score - 5),
        "uncertainty_analysis": "Analysis requires AI configuration. Based on keywords detected.",
        "experimentation_score": min(100, base_score),
        "experimentation_analysis": "Analysis requires AI configuration. Based on keywords detected.",
        "overall_score": base_score,
        "qualification_status": status,
        "qualification_narrative": f"Preliminary qualification based on keyword analysis. Configure OpenAI for full AI analysis.",
        "risk_flags": ["AI analysis unavailable - manual review recommended"],
        "suggested_evidence": ["Technical documentation", "Design specifications", "Test results"]
    }


async def analyze_employee_allocation_with_ai(openai_client, employee: RDEmployee, projects: List[RDProject]) -> dict:
    """
    Use OpenAI to analyze employee R&D time allocation based on their role and available projects.
    """
    if not openai_client:
        return _fallback_employee_allocation(employee)

    # Determine model/deployment name based on client type
    if settings.AZURE_OPENAI_ENDPOINT and settings.AZURE_OPENAI_API_KEY:
        model_name = settings.AZURE_OPENAI_DEPLOYMENT or "gpt-4-turbo"
    else:
        model_name = settings.OPENAI_CHAT_MODEL

    project_summaries = "\n".join([
        f"- {p.name}: {p.description[:200] if p.description else 'No description'}"
        for p in projects[:10]  # Limit to first 10 projects
    ])

    prompt = f"""Analyze this employee's likely R&D time allocation for tax credit purposes.

EMPLOYEE INFORMATION:
- Name: {employee.name}
- Job Title: {employee.title or 'Not specified'}
- Department: {employee.department or 'Not specified'}
- W-2 Wages: ${employee.w2_wages:,.2f}

COMPANY R&D PROJECTS:
{project_summaries or 'No project information available'}

Based on the employee's role and typical responsibilities for this job title, estimate:
1. What percentage of their time is likely spent on qualified R&D activities
2. What R&D activities they likely perform
3. Any concerns about the allocation

Respond with valid JSON only:
{{
    "qualified_time_percentage": <0-100>,
    "confidence": <0.0-1.0>,
    "activities": ["<list of likely R&D activities>"],
    "rationale": "<explanation of the allocation>",
    "risk_flags": ["<any concerns>"]
}}"""

    try:
        response = await openai_client.chat.completions.create(
            model=model_name,
            messages=[
                {"role": "system", "content": "You are an R&D tax credit specialist with expertise in employee time allocation studies. Be conservative and audit-defensible in your estimates."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.2,
            max_tokens=500
        )

        result_text = response.choices[0].message.content.strip()
        if result_text.startswith("```"):
            result_text = result_text.split("```")[1]
            if result_text.startswith("json"):
                result_text = result_text[4:]
        result = json.loads(result_text)
        logger.info(f"AI employee allocation successful using {model_name}")
        return result

    except Exception as e:
        logger.error(f"OpenAI employee allocation failed: {e}")
        return _fallback_employee_allocation(employee)


def _fallback_employee_allocation(employee: RDEmployee) -> dict:
    """Fallback allocation using title-based rules when OpenAI is unavailable."""
    title = (employee.title or "").lower()

    if any(t in title for t in ['engineer', 'developer', 'scientist', 'researcher']):
        pct = 75
        activities = ["Development", "Testing", "Design"]
    elif any(t in title for t in ['architect', 'lead', 'principal', 'senior']):
        pct = 65
        activities = ["Architecture", "Technical Leadership", "Code Review"]
    elif any(t in title for t in ['manager', 'director']):
        pct = 35
        activities = ["Project Planning", "Technical Oversight"]
    elif any(t in title for t in ['analyst', 'designer', 'qa', 'test']):
        pct = 50
        activities = ["Analysis", "Testing", "Documentation"]
    else:
        pct = 20
        activities = ["Support Activities"]

    return {
        "qualified_time_percentage": pct,
        "confidence": 0.6,
        "activities": activities,
        "rationale": f"Allocation based on job title '{employee.title}'. Configure OpenAI for detailed AI analysis.",
        "risk_flags": ["AI analysis unavailable - consider manual review"]
    }


# =============================================================================
# PAYROLL DATA UPLOAD
# =============================================================================

@router.post("/studies/{study_id}/upload/payroll")
async def upload_payroll_data(
    study_id: UUID,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db)
):
    """Upload and process payroll data from Excel file."""
    study = await db.get(RDStudy, study_id)
    if not study:
        raise HTTPException(status_code=404, detail="Study not found")

    # Read file content
    content = await file.read()

    try:
        # Parse Excel file
        workbook = openpyxl.load_workbook(io.BytesIO(content))
        sheet = workbook.active

        employees_created = 0
        employees_updated = 0

        # Expected columns: Name, Title, Department, W2 Wages, Qualified %
        headers = [cell.value for cell in sheet[1]]

        # Find column indices (flexible matching)
        name_col = next((i for i, h in enumerate(headers) if h and 'name' in str(h).lower()), 0)
        title_col = next((i for i, h in enumerate(headers) if h and 'title' in str(h).lower()), 1)
        dept_col = next((i for i, h in enumerate(headers) if h and 'dept' in str(h).lower()), 2)
        wages_col = next((i for i, h in enumerate(headers) if h and ('wage' in str(h).lower() or 'w2' in str(h).lower() or 'salary' in str(h).lower())), 3)
        pct_col = next((i for i, h in enumerate(headers) if h and ('%' in str(h) or 'percent' in str(h).lower() or 'qualified' in str(h).lower())), None)

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[name_col]:
                continue

            name = str(row[name_col]).strip()
            title = str(row[title_col]).strip() if row[title_col] else None
            department = str(row[dept_col]).strip() if row[dept_col] else None

            # Parse wages
            wages_value = row[wages_col]
            if isinstance(wages_value, str):
                wages_value = wages_value.replace('$', '').replace(',', '').strip()
            try:
                w2_wages = Decimal(str(wages_value)) if wages_value else Decimal('0')
            except:
                w2_wages = Decimal('0')

            # Parse qualified percentage
            qualified_pct = Decimal('0')
            if pct_col is not None and row[pct_col]:
                pct_value = row[pct_col]
                if isinstance(pct_value, str):
                    pct_value = pct_value.replace('%', '').strip()
                try:
                    qualified_pct = Decimal(str(pct_value))
                    if qualified_pct > 1:
                        qualified_pct = qualified_pct  # Already in percentage form
                    else:
                        qualified_pct = qualified_pct * 100  # Convert from decimal
                except:
                    qualified_pct = Decimal('0')

            # Calculate qualified wages
            qualified_wages = w2_wages * (qualified_pct / 100)

            # Check if employee exists
            existing_result = await db.execute(
                select(RDEmployee).where(
                    RDEmployee.study_id == study_id,
                    RDEmployee.name == name
                )
            )
            existing = existing_result.scalars().first()

            if existing:
                # Update existing employee
                existing.title = title or existing.title
                existing.department = department or existing.department
                existing.w2_wages = w2_wages
                existing.total_wages = w2_wages
                existing.qualified_time_percentage = qualified_pct
                existing.qualified_wages = qualified_wages
                existing.updated_at = datetime.utcnow()
                employees_updated += 1
            else:
                # Create new employee
                employee = RDEmployee(
                    study_id=study_id,
                    name=name,
                    title=title,
                    department=department,
                    w2_wages=w2_wages,
                    total_wages=w2_wages,
                    qualified_time_percentage=qualified_pct,
                    qualified_wages=qualified_wages,
                    qualified_time_source="payroll_import"
                )
                db.add(employee)
                employees_created += 1

        # Save document record
        document = RDDocument(
            study_id=study_id,
            filename=file.filename,
            file_type=file.content_type,
            document_type="payroll",
            file_size=len(content),
            processing_status="completed",
            storage_path=f"uploads/{study_id}/{file.filename}"
        )
        db.add(document)

        await db.commit()

        return {
            "message": "Payroll data processed successfully",
            "employees_created": employees_created,
            "employees_updated": employees_updated,
            "total_processed": employees_created + employees_updated
        }

    except Exception as e:
        logger.error(f"Error processing payroll file: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")


# =============================================================================
# PROJECT DATA UPLOAD
# =============================================================================

@router.post("/studies/{study_id}/upload/projects")
async def upload_project_data(
    study_id: UUID,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db)
):
    """Upload and process project data from Excel file."""
    study = await db.get(RDStudy, study_id)
    if not study:
        raise HTTPException(status_code=404, detail="Study not found")

    content = await file.read()

    try:
        workbook = openpyxl.load_workbook(io.BytesIO(content))
        sheet = workbook.active

        projects_created = 0
        headers = [cell.value for cell in sheet[1]]

        # Find column indices
        name_col = next((i for i, h in enumerate(headers) if h and 'name' in str(h).lower()), 0)
        desc_col = next((i for i, h in enumerate(headers) if h and 'desc' in str(h).lower()), 1)
        dept_col = next((i for i, h in enumerate(headers) if h and 'dept' in str(h).lower()), 2)

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[name_col]:
                continue

            project = RDProject(
                study_id=study_id,
                name=str(row[name_col]).strip(),
                description=str(row[desc_col]).strip() if desc_col < len(row) and row[desc_col] else None,
                department=str(row[dept_col]).strip() if dept_col < len(row) and row[dept_col] else None,
                qualification_status=QualificationStatus.PENDING
            )
            db.add(project)
            projects_created += 1

        # Save document record
        document = RDDocument(
            study_id=study_id,
            filename=file.filename,
            file_type=file.content_type,
            document_type="project_list",
            file_size=len(content),
            processing_status="completed",
            storage_path=f"uploads/{study_id}/{file.filename}"
        )
        db.add(document)

        await db.commit()

        return {
            "message": "Project data processed successfully",
            "projects_created": projects_created
        }

    except Exception as e:
        logger.error(f"Error processing project file: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")


# =============================================================================
# QRE DATA UPLOAD
# =============================================================================

@router.post("/studies/{study_id}/upload/expenses")
async def upload_expense_data(
    study_id: UUID,
    file: UploadFile = File(...),
    expense_type: str = Form(default="supplies"),  # supplies, contract_research
    db: AsyncSession = Depends(get_db)
):
    """Upload and process expense data (supplies, contract research) from Excel file."""
    study = await db.get(RDStudy, study_id)
    if not study:
        raise HTTPException(status_code=404, detail="Study not found")

    content = await file.read()

    try:
        workbook = openpyxl.load_workbook(io.BytesIO(content))
        sheet = workbook.active

        qres_created = 0
        headers = [cell.value for cell in sheet[1]]

        # Find column indices
        desc_col = next((i for i, h in enumerate(headers) if h and 'desc' in str(h).lower()), 0)
        vendor_col = next((i for i, h in enumerate(headers) if h and 'vendor' in str(h).lower()), 1)
        amount_col = next((i for i, h in enumerate(headers) if h and 'amount' in str(h).lower()), 2)
        gl_col = next((i for i, h in enumerate(headers) if h and ('gl' in str(h).lower() or 'account' in str(h).lower())), None)

        category = QRECategory.SUPPLIES if expense_type == "supplies" else QRECategory.CONTRACT_RESEARCH

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[amount_col]:
                continue

            # Parse amount
            amount_value = row[amount_col]
            if isinstance(amount_value, str):
                amount_value = amount_value.replace('$', '').replace(',', '').strip()
            try:
                gross_amount = Decimal(str(amount_value))
            except:
                continue

            # For supplies, 100% qualified
            # For contract research, 65% for non-qualified orgs, 75% for qualified orgs
            qualified_pct = Decimal('100') if expense_type == "supplies" else Decimal('65')
            qualified_amount = gross_amount * (qualified_pct / 100)

            qre = QualifiedResearchExpense(
                study_id=study_id,
                category=category,
                description=str(row[desc_col]).strip() if row[desc_col] else expense_type,
                supply_description=str(row[desc_col]).strip() if expense_type == "supplies" and row[desc_col] else None,
                supply_vendor=str(row[vendor_col]).strip() if expense_type == "supplies" and vendor_col < len(row) and row[vendor_col] else None,
                contractor_name=str(row[vendor_col]).strip() if expense_type == "contract_research" and vendor_col < len(row) and row[vendor_col] else None,
                gross_amount=gross_amount,
                qualified_percentage=qualified_pct,
                qualified_amount=qualified_amount,
                gl_account=str(row[gl_col]).strip() if gl_col is not None and gl_col < len(row) and row[gl_col] else None,
            )
            db.add(qre)
            qres_created += 1

        # Save document record
        document = RDDocument(
            study_id=study_id,
            filename=file.filename,
            file_type=file.content_type,
            document_type="general_ledger" if expense_type == "supplies" else "contractor_invoices",
            file_size=len(content),
            processing_status="completed",
            storage_path=f"uploads/{study_id}/{file.filename}"
        )
        db.add(document)

        await db.commit()

        return {
            "message": f"{expense_type.replace('_', ' ').title()} data processed successfully",
            "qres_created": qres_created
        }

    except Exception as e:
        logger.error(f"Error processing expense file: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")


# =============================================================================
# AI STUDY COMPLETION
# =============================================================================

@router.post("/studies/{study_id}/ai/complete-study")
async def ai_complete_study(
    study_id: UUID,
    request: Request,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db)
):
    """
    AI-powered study completion.
    Analyzes all data and generates:
    - Project qualifications
    - Employee allocations
    - QRE calculations
    - Credit calculations
    """
    # Get OpenAI client from app state
    openai_client = getattr(request.app.state, 'openai_client', None)

    study = await db.get(RDStudy, study_id)
    if not study:
        raise HTTPException(status_code=404, detail="Study not found")

    # Get all employees
    employees_result = await db.execute(
        select(RDEmployee).where(RDEmployee.study_id == study_id)
    )
    employees = employees_result.scalars().all()

    # Get all projects
    projects_result = await db.execute(
        select(RDProject).where(RDProject.study_id == study_id)
    )
    projects = projects_result.scalars().all()

    # Get all QREs
    qres_result = await db.execute(
        select(QualifiedResearchExpense).where(QualifiedResearchExpense.study_id == study_id)
    )
    qres = qres_result.scalars().all()

    # AI ANALYSIS: Qualify projects using real AI
    ai_used = openai_client is not None
    for project in projects:
        if project.qualification_status == QualificationStatus.PENDING:
            # Use real AI analysis for project qualification
            qual_result = await qualify_project_with_ai(openai_client, project)

            # Apply results to project
            project.permitted_purpose_score = Decimal(str(qual_result.get("permitted_purpose_score", 70)))
            project.permitted_purpose_narrative = qual_result.get("permitted_purpose_analysis", "")
            project.technological_nature_score = Decimal(str(qual_result.get("technological_nature_score", 70)))
            project.technological_nature_narrative = qual_result.get("technological_nature_analysis", "")
            project.uncertainty_score = Decimal(str(qual_result.get("uncertainty_score", 65)))
            project.uncertainty_narrative = qual_result.get("uncertainty_analysis", "")
            project.experimentation_score = Decimal(str(qual_result.get("experimentation_score", 70)))
            project.experimentation_narrative = qual_result.get("experimentation_analysis", "")
            project.overall_score = Decimal(str(qual_result.get("overall_score", 70)))
            project.qualification_narrative = qual_result.get("qualification_narrative", "")
            project.risk_flags = qual_result.get("risk_flags", [])
            project.ai_qualification_analysis = {
                "suggested_evidence": qual_result.get("suggested_evidence", []),
                "ai_analysis_used": ai_used,
                "analyzed_at": datetime.utcnow().isoformat()
            }

            # Set qualification status
            status_str = qual_result.get("qualification_status", "needs_review")
            if status_str == "qualified":
                project.qualification_status = QualificationStatus.QUALIFIED
            elif status_str == "not_qualified":
                project.qualification_status = QualificationStatus.NOT_QUALIFIED
            else:
                project.qualification_status = QualificationStatus.NEEDS_REVIEW

    # AI ANALYSIS: Adjust employee allocations using real AI
    for employee in employees:
        # If no qualified percentage set, use AI to estimate
        if not employee.qualified_time_percentage or employee.qualified_time_percentage == 0:
            alloc_result = await analyze_employee_allocation_with_ai(openai_client, employee, projects)

            employee.qualified_time_percentage = Decimal(str(alloc_result.get("qualified_time_percentage", 25)))
            employee.qualified_time_confidence = alloc_result.get("confidence", 0.5)
            employee.risk_flags = alloc_result.get("risk_flags", [])
            employee.ai_allocation_analysis = {
                "activities": alloc_result.get("activities", []),
                "rationale": alloc_result.get("rationale", ""),
                "ai_analysis_used": ai_used,
                "analyzed_at": datetime.utcnow().isoformat()
            }

            # Calculate qualified wages
            if employee.w2_wages:
                employee.qualified_wages = employee.w2_wages * (employee.qualified_time_percentage / 100)

    # Calculate totals
    total_wage_qre = sum(e.qualified_wages or Decimal('0') for e in employees)
    total_supply_qre = sum(q.qualified_amount or Decimal('0') for q in qres if q.category == QRECategory.SUPPLIES)
    total_contract_qre = sum(q.qualified_amount or Decimal('0') for q in qres if q.category == QRECategory.CONTRACT_RESEARCH)
    total_qre = total_wage_qre + total_supply_qre + total_contract_qre

    # Calculate credits
    # ASC method: 14% of (QRE - 50% of average prior 3 years QRE)
    # For first year/no history: 6% of QRE
    asc_credit = total_qre * Decimal('0.06')  # First year rate

    # Regular method: 20% of (QRE - Base)
    # Simplified: 20% of 50% of QRE for startups
    regular_credit = total_qre * Decimal('0.5') * Decimal('0.20')

    # Apply Section 280C reduction (21% corporate tax rate)
    asc_credit_final = asc_credit * Decimal('0.79')
    regular_credit_final = regular_credit * Decimal('0.79')

    # Select better method
    selected_method = CreditMethod.ASC if asc_credit_final >= regular_credit_final else CreditMethod.REGULAR
    final_credit = max(asc_credit_final, regular_credit_final)

    # Calculate state credits based on employee states
    from ..engines.calculation_engine import CalculationEngine
    from ..engines.rules_engine import RulesEngine

    # Get unique states from employees
    employee_states = set()
    for emp in employees:
        emp_state = getattr(emp, 'state', None)
        if emp_state and len(emp_state) == 2:
            employee_states.add(emp_state.upper())

    # Calculate state credits
    total_state_credits = Decimal('0')
    state_results = {}

    if employee_states:
        rules_engine = RulesEngine()
        calc_engine = CalculationEngine(rules_engine)

        for state_code in employee_states:
            state_rules = rules_engine.get_state_rules(state_code)
            if state_rules and state_rules.has_rd_credit:
                # Calculate state QRE allocation based on employees in that state
                state_wage_qre = sum(
                    e.qualified_wages or Decimal('0')
                    for e in employees
                    if getattr(e, 'state', '').upper() == state_code
                )

                state_qre_data = {
                    "wages": state_wage_qre,
                    "supplies": Decimal('0'),  # Allocate supplies by state if tracked
                    "contract_research": Decimal('0')
                }

                state_result = calc_engine.calculate_state_credit(
                    state_code=state_code,
                    tax_year=study.tax_year or 2024,
                    qre_data=state_qre_data,
                    federal_qre=total_qre,
                    federal_base_amount=Decimal('0')
                )

                if state_result:
                    state_results[state_code] = {
                        "state_name": state_result.state_name,
                        "credit_rate": float(state_result.credit_rate),
                        "credit_type": state_result.credit_type,
                        "state_qre": float(state_result.state_qre),
                        "final_credit": float(state_result.final_credit),
                        "carryforward_years": state_result.carryforward_years,
                        "state_form": state_result.state_form
                    }
                    total_state_credits += state_result.final_credit

    total_credits = final_credit + total_state_credits

    # Update study
    study.total_qre = total_qre
    study.qre_wages = total_wage_qre
    study.qre_supplies = total_supply_qre
    study.qre_contract = total_contract_qre
    study.federal_credit_regular = regular_credit_final
    study.federal_credit_asc = asc_credit_final
    study.federal_credit_final = final_credit
    study.selected_method = selected_method
    study.total_credits = total_credits
    study.status = StudyStatus.CPA_APPROVAL
    study.updated_at = datetime.utcnow()

    # Store state results in ai_analysis
    if not study.ai_suggested_areas:
        study.ai_suggested_areas = {}
    study.ai_suggested_areas["state_credits"] = state_results
    study.ai_suggested_areas["total_state_credits"] = float(total_state_credits)

    # Create calculation record
    # credit_rate is required (NOT NULL): 0.14 for ASC, 0.20 for Regular
    credit_rate = Decimal('0.14') if selected_method == CreditMethod.ASC else Decimal('0.20')
    calculation = RDCalculation(
        study_id=study_id,
        calculation_type="federal_asc" if selected_method == CreditMethod.ASC else "federal_regular",
        total_qre=total_qre,
        calculated_credit=final_credit,
        credit_rate=credit_rate,
        is_final=False,
        calculation_steps=[
            {"step": 1, "description": "Total Wage QRE", "value": str(total_wage_qre)},
            {"step": 2, "description": "Total Supply QRE", "value": str(total_supply_qre)},
            {"step": 3, "description": "Total Contract QRE", "value": str(total_contract_qre)},
            {"step": 4, "description": "Total QRE", "value": str(total_qre)},
            {"step": 5, "description": "Credit Rate", "value": "14%" if selected_method == CreditMethod.ASC else "20%"},
            {"step": 6, "description": "Tentative Credit", "value": str(asc_credit if selected_method == CreditMethod.ASC else regular_credit)},
            {"step": 7, "description": "Section 280C Reduction (21%)", "value": "Applied"},
            {"step": 8, "description": "Final Credit", "value": str(final_credit)},
        ]
    )
    db.add(calculation)

    await db.commit()

    return {
        "message": "AI study completion successful",
        "study_id": str(study_id),
        "ai_enabled": ai_used,
        "projects_analyzed": len(projects),
        "employees_analyzed": len(employees),
        "qres_analyzed": len(qres),
        "results": {
            "total_qre": float(total_qre),
            "wage_qre": float(total_wage_qre),
            "supply_qre": float(total_supply_qre),
            "contract_qre": float(total_contract_qre),
            "federal_credit_regular": float(regular_credit_final),
            "federal_credit_asc": float(asc_credit_final),
            "selected_method": selected_method.value,
            "final_credit": float(final_credit),
        },
        "status": "cpa_review",
        "ai_analysis": {
            "openai_configured": ai_used,
            "model": settings.OPENAI_CHAT_MODEL if ai_used else None,
            "note": "Full AI analysis completed" if ai_used else "Fallback analysis used - configure OPENAI_API_KEY for production"
        }
    }


# =============================================================================
# QUICK STUDY WIZARD
# =============================================================================

# =============================================================================
# 4-PART TEST QUESTIONNAIRE
# =============================================================================

@router.get("/studies/{study_id}/projects/{project_id}/questionnaire")
async def get_project_questionnaire(
    study_id: UUID,
    project_id: UUID,
    db: AsyncSession = Depends(get_db)
):
    """Get saved questionnaire answers for a project."""
    project = await db.get(RDProject, project_id)
    if not project or project.study_id != study_id:
        raise HTTPException(status_code=404, detail="Project not found")

    # Return stored questionnaire answers from ai_qualification_analysis
    answers = {}
    if project.ai_qualification_analysis and isinstance(project.ai_qualification_analysis, dict):
        answers = project.ai_qualification_analysis.get("questionnaire_answers", {})

    return {
        "project_id": str(project_id),
        "project_name": project.name,
        "answers": answers,
        "has_answers": bool(answers)
    }


@router.post("/studies/{study_id}/projects/{project_id}/questionnaire")
async def save_project_questionnaire(
    study_id: UUID,
    project_id: UUID,
    answers: dict,
    db: AsyncSession = Depends(get_db)
):
    """
    Save 4-Part Test questionnaire answers for a project.
    The questionnaire covers:
    - Part 1: Permitted Purpose (business component improvement)
    - Part 2: Technological in Nature (relies on hard sciences)
    - Part 3: Elimination of Uncertainty (capability, method, design)
    - Part 4: Process of Experimentation (systematic evaluation)
    """
    project = await db.get(RDProject, project_id)
    if not project or project.study_id != study_id:
        raise HTTPException(status_code=404, detail="Project not found")

    # Validate and score answers
    part_scores = {}

    # Part 1: Permitted Purpose
    part1 = answers.get("part1", {})
    part1_score = 0
    if part1.get("improves_business_component"):
        part1_score += 40
    if part1.get("business_component_type"):
        part1_score += 20
    if part1.get("improvement_type") in ["new_development", "improvement", "capability"]:
        part1_score += 40
    part_scores["permitted_purpose"] = min(100, part1_score)

    # Part 2: Technological in Nature
    part2 = answers.get("part2", {})
    part2_score = 0
    sciences = part2.get("hard_sciences", [])
    if sciences:
        part2_score += 50
        if len(sciences) > 1:
            part2_score += 20
    if part2.get("relies_on_principles"):
        part2_score += 30
    part_scores["technological_nature"] = min(100, part2_score)

    # Part 3: Elimination of Uncertainty
    part3 = answers.get("part3", {})
    part3_score = 0
    uncertainties = part3.get("uncertainty_types", [])
    if "capability" in uncertainties:
        part3_score += 35
    if "method" in uncertainties:
        part3_score += 35
    if "design" in uncertainties:
        part3_score += 30
    if part3.get("uncertainty_description"):
        part3_score += 10
    part_scores["uncertainty"] = min(100, part3_score)

    # Part 4: Process of Experimentation
    part4 = answers.get("part4", {})
    part4_score = 0
    methods = part4.get("experimentation_methods", [])
    if "modeling" in methods:
        part4_score += 25
    if "simulation" in methods:
        part4_score += 25
    if "systematic_trial_error" in methods:
        part4_score += 30
    if "other_evaluation" in methods:
        part4_score += 20
    if part4.get("evaluated_alternatives"):
        part4_score += 20
    part_scores["experimentation"] = min(100, part4_score)

    # Overall score
    overall_score = sum(part_scores.values()) / 4

    # Determine qualification status
    if overall_score >= 75:
        qual_status = "qualified"
    elif overall_score >= 50:
        qual_status = "needs_review"
    else:
        qual_status = "not_qualified"

    # Store answers and scores
    if not project.ai_qualification_analysis:
        project.ai_qualification_analysis = {}

    project.ai_qualification_analysis["questionnaire_answers"] = answers
    project.ai_qualification_analysis["questionnaire_scores"] = part_scores
    project.ai_qualification_analysis["questionnaire_completed_at"] = datetime.utcnow().isoformat()

    # Update project scores
    project.permitted_purpose_score = Decimal(str(part_scores["permitted_purpose"]))
    project.technological_nature_score = Decimal(str(part_scores["technological_nature"]))
    project.uncertainty_score = Decimal(str(part_scores["uncertainty"]))
    project.experimentation_score = Decimal(str(part_scores["experimentation"]))
    project.overall_score = Decimal(str(overall_score))

    # Set qualification status based on score
    if qual_status == "qualified":
        project.qualification_status = QualificationStatus.QUALIFIED
    elif qual_status == "not_qualified":
        project.qualification_status = QualificationStatus.NOT_QUALIFIED
    else:
        project.qualification_status = QualificationStatus.NEEDS_REVIEW

    project.updated_at = datetime.utcnow()
    await db.commit()

    return {
        "success": True,
        "project_id": str(project_id),
        "scores": part_scores,
        "overall_score": overall_score,
        "qualification_status": qual_status,
        "message": "Questionnaire saved successfully"
    }


@router.post("/studies/{study_id}/projects/{project_id}/generate-narrative")
async def generate_project_narrative(
    study_id: UUID,
    project_id: UUID,
    request: Request,
    payload: dict,
    db: AsyncSession = Depends(get_db)
):
    """
    Generate AI-powered technical narrative for the 4-Part Test.
    Uses questionnaire answers and project details to create audit-defensible documentation.
    """
    openai_client = getattr(request.app.state, 'openai_client', None)

    project = await db.get(RDProject, project_id)
    if not project or project.study_id != study_id:
        raise HTTPException(status_code=404, detail="Project not found")

    questionnaire_answers = payload.get("questionnaire_answers", {})
    if not questionnaire_answers:
        # Try to get from stored data
        if project.ai_qualification_analysis:
            questionnaire_answers = project.ai_qualification_analysis.get("questionnaire_answers", {})

    if not questionnaire_answers:
        raise HTTPException(
            status_code=400,
            detail="No questionnaire answers provided. Please complete the 4-Part Test questionnaire first."
        )

    # Build comprehensive project context
    part1 = questionnaire_answers.get("part1", {})
    part2 = questionnaire_answers.get("part2", {})
    part3 = questionnaire_answers.get("part3", {})
    part4 = questionnaire_answers.get("part4", {})

    if openai_client:
        try:
            # Determine model
            if settings.AZURE_OPENAI_ENDPOINT and settings.AZURE_OPENAI_API_KEY:
                model_name = settings.AZURE_OPENAI_DEPLOYMENT or "gpt-4-turbo"
            else:
                model_name = settings.OPENAI_CHAT_MODEL

            prompt = f"""Generate a comprehensive technical narrative for this R&D project's 4-Part Test qualification.
This narrative should be audit-defensible and suitable for IRS Form 6765 documentation.

PROJECT INFORMATION:
- Name: {project.name}
- Description: {project.description or 'Not provided'}
- Business Component: {project.business_component or part1.get('business_component_type', 'Not specified')}

QUESTIONNAIRE RESPONSES:

PART 1 - PERMITTED PURPOSE:
- Improves Business Component: {part1.get('improves_business_component', False)}
- Component Type: {part1.get('business_component_type', 'Not specified')}
- Improvement Type: {part1.get('improvement_type', 'Not specified')}
- Description: {part1.get('improvement_description', 'Not provided')}

PART 2 - TECHNOLOGICAL IN NATURE:
- Hard Sciences Used: {', '.join(part2.get('hard_sciences', [])) or 'None specified'}
- Relies on Scientific Principles: {part2.get('relies_on_principles', False)}
- Technical Description: {part2.get('technical_description', 'Not provided')}

PART 3 - ELIMINATION OF UNCERTAINTY:
- Uncertainty Types: {', '.join(part3.get('uncertainty_types', [])) or 'None specified'}
- Uncertainty Description: {part3.get('uncertainty_description', 'Not provided')}
- What Was Unknown: {part3.get('what_was_unknown', 'Not provided')}

PART 4 - PROCESS OF EXPERIMENTATION:
- Methods Used: {', '.join(part4.get('experimentation_methods', [])) or 'None specified'}
- Evaluated Alternatives: {part4.get('evaluated_alternatives', False)}
- Experimentation Description: {part4.get('experimentation_description', 'Not provided')}

Generate a JSON response with narratives for each part of the test:
{{
    "permitted_purpose_narrative": "<detailed narrative explaining how this project meets the Permitted Purpose test>",
    "technological_nature_narrative": "<detailed narrative explaining the technological nature and hard sciences involved>",
    "uncertainty_narrative": "<detailed narrative describing the technical uncertainties faced>",
    "experimentation_narrative": "<detailed narrative describing the systematic process of experimentation>",
    "executive_summary": "<2-3 sentence summary of why this project qualifies for R&D tax credit>",
    "documentation_recommendations": ["<list of additional documentation that would strengthen the case>"]
}}"""

            response = await openai_client.chat.completions.create(
                model=model_name,
                messages=[
                    {
                        "role": "system",
                        "content": "You are an expert R&D tax credit consultant with deep knowledge of IRC Section 41. Generate professional, audit-defensible technical narratives that clearly demonstrate qualification under the 4-Part Test. Use specific technical language and reference the actual project activities."
                    },
                    {"role": "user", "content": prompt}
                ],
                temperature=0.3,
                max_tokens=2500
            )

            result_text = response.choices[0].message.content.strip()
            if result_text.startswith("```"):
                result_text = result_text.split("```")[1]
                if result_text.startswith("json"):
                    result_text = result_text[4:]

            narratives = json.loads(result_text)
            ai_used = True
            logger.info(f"AI narrative generation successful for project {project_id}")

        except Exception as e:
            logger.error(f"AI narrative generation failed: {e}")
            narratives = _fallback_generate_narratives(project, questionnaire_answers)
            ai_used = False
    else:
        narratives = _fallback_generate_narratives(project, questionnaire_answers)
        ai_used = False

    # Store narratives in project
    project.permitted_purpose_narrative = narratives.get("permitted_purpose_narrative", "")
    project.technological_nature_narrative = narratives.get("technological_nature_narrative", "")
    project.uncertainty_narrative = narratives.get("uncertainty_narrative", "")
    project.experimentation_narrative = narratives.get("experimentation_narrative", "")
    project.qualification_narrative = narratives.get("executive_summary", "")

    if not project.ai_qualification_analysis:
        project.ai_qualification_analysis = {}
    project.ai_qualification_analysis["ai_narratives"] = narratives
    project.ai_qualification_analysis["narratives_generated_at"] = datetime.utcnow().isoformat()
    project.ai_qualification_analysis["ai_used"] = ai_used

    project.updated_at = datetime.utcnow()
    await db.commit()

    return {
        "success": True,
        "project_id": str(project_id),
        "narratives": narratives,
        "ai_used": ai_used,
        "message": "Technical narratives generated successfully"
    }


def _fallback_generate_narratives(project: RDProject, answers: dict) -> dict:
    """Generate fallback narratives when AI is unavailable."""
    part1 = answers.get("part1", {})
    part2 = answers.get("part2", {})
    part3 = answers.get("part3", {})
    part4 = answers.get("part4", {})

    component_type = part1.get("business_component_type", "product")
    sciences = part2.get("hard_sciences", ["engineering"])
    uncertainties = part3.get("uncertainty_types", ["capability"])
    methods = part4.get("experimentation_methods", ["systematic_trial_error"])

    return {
        "permitted_purpose_narrative": f"The {project.name} project was undertaken with the intent to develop or improve a {component_type} for use in the taxpayer's trade or business. {part1.get('improvement_description', 'The project involved substantial development activities aimed at improving functionality, performance, or reliability.')}",

        "technological_nature_narrative": f"This project's activities fundamentally rely on principles of {', '.join(sciences)}. The technical team applied established scientific methodologies and engineering principles to achieve the project objectives. {part2.get('technical_description', 'Development required specialized technical knowledge and expertise.')}",

        "uncertainty_narrative": f"At the outset of this project, the development team faced significant technical uncertainties regarding {', '.join(uncertainties)}. {part3.get('uncertainty_description', 'These uncertainties could not be resolved through conventional means and required systematic technical investigation.')} The capability to achieve the desired result was not known prior to the commencement of research activities.",

        "experimentation_narrative": f"To resolve the technical uncertainties, the project team employed a systematic process of experimentation including {', '.join([m.replace('_', ' ') for m in methods])}. {part4.get('experimentation_description', 'Multiple approaches were evaluated and tested to determine the optimal solution.')} This iterative process of hypothesis, testing, and refinement is consistent with the requirements of IRC Section 41.",

        "executive_summary": f"The {project.name} project qualifies for the R&D tax credit under IRC Section 41 as it involves permitted purpose activities (developing/improving a {component_type}), relies on {sciences[0] if sciences else 'hard science'} principles, addresses technical uncertainties regarding {uncertainties[0] if uncertainties else 'capability'}, and employs systematic experimentation to resolve those uncertainties.",

        "documentation_recommendations": [
            "Technical design documents and specifications",
            "Meeting notes discussing technical challenges",
            "Test results and iteration records",
            "Email communications regarding technical decisions",
            "Employee time tracking for R&D activities"
        ]
    }


@router.post("/studies/{study_id}/wizard/quick-complete")
async def quick_complete_wizard(
    study_id: UUID,
    payroll_file: Optional[UploadFile] = File(None),
    projects_file: Optional[UploadFile] = File(None),
    supplies_file: Optional[UploadFile] = File(None),
    db: AsyncSession = Depends(get_db)
):
    """
    Quick wizard to upload all data and complete study in one step.
    """
    study = await db.get(RDStudy, study_id)
    if not study:
        raise HTTPException(status_code=404, detail="Study not found")

    results = {
        "payroll": None,
        "projects": None,
        "supplies": None,
        "study_completion": None
    }

    # Process payroll file if provided
    if payroll_file:
        try:
            content = await payroll_file.read()
            workbook = openpyxl.load_workbook(io.BytesIO(content))
            sheet = workbook.active

            employees_created = 0
            headers = [cell.value for cell in sheet[1]]
            name_col = next((i for i, h in enumerate(headers) if h and 'name' in str(h).lower()), 0)
            wages_col = next((i for i, h in enumerate(headers) if h and ('wage' in str(h).lower() or 'salary' in str(h).lower())), 3)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[name_col]:
                    continue

                wages_value = row[wages_col]
                if isinstance(wages_value, str):
                    wages_value = wages_value.replace('$', '').replace(',', '').strip()
                try:
                    w2_wages = Decimal(str(wages_value)) if wages_value else Decimal('0')
                except:
                    w2_wages = Decimal('0')

                employee = RDEmployee(
                    study_id=study_id,
                    name=str(row[name_col]).strip(),
                    w2_wages=w2_wages,
                    total_wages=w2_wages,
                    qualified_time_percentage=Decimal('50'),  # Default 50%
                    qualified_wages=w2_wages * Decimal('0.5'),
                    qualified_time_source="wizard_import"
                )
                db.add(employee)
                employees_created += 1

            results["payroll"] = {"employees_created": employees_created}
        except Exception as e:
            results["payroll"] = {"error": str(e)}

    # Process projects file if provided
    if projects_file:
        try:
            content = await projects_file.read()
            workbook = openpyxl.load_workbook(io.BytesIO(content))
            sheet = workbook.active

            projects_created = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue

                project = RDProject(
                    study_id=study_id,
                    name=str(row[0]).strip(),
                    description=str(row[1]).strip() if len(row) > 1 and row[1] else None,
                    qualification_status=QualificationStatus.PENDING
                )
                db.add(project)
                projects_created += 1

            results["projects"] = {"projects_created": projects_created}
        except Exception as e:
            results["projects"] = {"error": str(e)}

    # Process supplies file if provided
    if supplies_file:
        try:
            content = await supplies_file.read()
            workbook = openpyxl.load_workbook(io.BytesIO(content))
            sheet = workbook.active

            qres_created = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue

                amount_value = row[1] if len(row) > 1 else 0
                if isinstance(amount_value, str):
                    amount_value = amount_value.replace('$', '').replace(',', '').strip()
                try:
                    gross_amount = Decimal(str(amount_value))
                except:
                    continue

                qre = QualifiedResearchExpense(
                    study_id=study_id,
                    category=QRECategory.SUPPLIES,
                    description=str(row[0]).strip(),
                    gross_amount=gross_amount,
                    qualified_percentage=Decimal('100'),
                    qualified_amount=gross_amount,
                )
                db.add(qre)
                qres_created += 1

            results["supplies"] = {"qres_created": qres_created}
        except Exception as e:
            results["supplies"] = {"error": str(e)}

    await db.commit()

    # Now run AI study completion
    # (reusing logic from ai_complete_study)
    employees_result = await db.execute(
        select(RDEmployee).where(RDEmployee.study_id == study_id)
    )
    employees = employees_result.scalars().all()

    qres_result = await db.execute(
        select(QualifiedResearchExpense).where(QualifiedResearchExpense.study_id == study_id)
    )
    qres = qres_result.scalars().all()

    # Calculate totals
    total_wage_qre = sum(e.qualified_wages or Decimal('0') for e in employees)
    total_supply_qre = sum(q.qualified_amount or Decimal('0') for q in qres if q.category == QRECategory.SUPPLIES)
    total_qre = total_wage_qre + total_supply_qre

    # ASC calculation (first year: 6% of QRE)
    asc_credit = total_qre * Decimal('0.06')
    final_credit = asc_credit * Decimal('0.79')  # After 280C reduction

    # Update study totals (only total_qre exists on model)
    study.total_qre = total_qre
    study.federal_credit_asc = final_credit
    study.federal_credit_final = final_credit
    study.selected_method = CreditMethod.ASC
    study.total_credits = final_credit
    study.status = StudyStatus.CPA_APPROVAL

    await db.commit()

    results["study_completion"] = {
        "total_qre": float(total_qre),
        "federal_credit": float(final_credit),
        "status": "cpa_review"
    }

    return {
        "message": "Quick wizard completed",
        "study_id": str(study_id),
        "results": results
    }
