"""
Excel Workbook Generator

Generates comprehensive Excel workbooks for R&D tax credit studies.
Uses openpyxl for real Excel file generation.
"""

import io
import logging
from typing import Dict, List, Optional, Any
from datetime import datetime
from decimal import Decimal
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import BarChart, Reference, PieChart

logger = logging.getLogger(__name__)


class ExcelWorkbookGenerator:
    """
    Generates Excel workbooks for R&D tax credit studies.

    Workbook includes:
    - Summary Dashboard
    - Raw Imports (read-only)
    - Normalized Data
    - QRE Schedules
    - Federal Regular Calculation
    - Federal ASC Calculation
    - State Tabs
    - GL/Payroll Reconciliations
    - Sanity Checks
    - Form 6765 Data
    """

    # Style definitions
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    SUBHEADER_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    SUBHEADER_FONT = Font(bold=True, size=11)
    TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
    SECTION_FONT = Font(bold=True, size=12, color="1F4E79")
    CURRENCY_FORMAT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    PERCENTAGE_FORMAT = '0.00%'
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    TOTALS_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    TOTALS_FONT = Font(bold=True)

    def __init__(self, config: Optional[Dict] = None):
        self.config = config or {}

    def generate_workbook(
        self,
        study_data: Dict[str, Any],
        projects: List[Dict[str, Any]],
        employees: List[Dict[str, Any]],
        qres: List[Dict[str, Any]],
        calculation_result: Dict[str, Any],
        raw_data: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """
        Generate complete Excel workbook.

        Returns workbook content as bytes.
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create sheets
        self._create_summary_sheet(wb, study_data, calculation_result, projects, employees)
        self._create_qre_summary_sheet(wb, qres, calculation_result)
        self._create_employees_sheet(wb, employees)
        self._create_projects_sheet(wb, projects)
        self._create_wage_qre_sheet(wb, employees, qres)
        self._create_supply_qre_sheet(wb, qres)
        self._create_contract_qre_sheet(wb, qres)
        self._create_federal_regular_sheet(wb, calculation_result)
        self._create_federal_asc_sheet(wb, calculation_result)

        # Add state sheets
        for state_code, state_result in calculation_result.get("state_results", {}).items():
            self._create_state_sheet(wb, state_code, state_result)

        self._create_reconciliation_sheet(wb, study_data, qres, employees)
        self._create_sanity_checks_sheet(wb, study_data, calculation_result)
        self._create_form_6765_sheet(wb, study_data, calculation_result)

        if raw_data:
            self._create_raw_imports_sheet(wb, raw_data)

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _apply_header_style(self, ws: Worksheet, row: int, start_col: int, end_col: int):
        """Apply header styling to a row."""
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.THIN_BORDER

    def _apply_data_style(self, cell, is_currency: bool = False, is_percentage: bool = False):
        """Apply data cell styling."""
        cell.border = self.THIN_BORDER
        cell.alignment = Alignment(horizontal='right' if is_currency or is_percentage else 'left')
        if is_currency:
            cell.number_format = self.CURRENCY_FORMAT
        elif is_percentage:
            cell.number_format = self.PERCENTAGE_FORMAT

    def _set_column_widths(self, ws: Worksheet, widths: Dict[str, int]):
        """Set column widths."""
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width

    def _create_summary_sheet(
        self,
        wb: Workbook,
        study_data: Dict,
        calculation_result: Dict,
        projects: List[Dict],
        employees: List[Dict]
    ):
        """Create summary dashboard sheet."""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws['A1'] = "R&D Tax Credit Study Summary"
        ws['A1'].font = Font(bold=True, size=18, color="1F4E79")
        ws.merge_cells('A1:F1')

        # Study Information
        row = 3
        ws[f'A{row}'] = "Study Information"
        ws[f'A{row}'].font = self.SECTION_FONT
        row += 1

        info_data = [
            ("Entity Name", study_data.get("entity_name", "")),
            ("Tax Year", study_data.get("tax_year", "")),
            ("EIN", study_data.get("ein", "")),
            ("Entity Type", study_data.get("entity_type", "").replace("_", " ").title()),
            ("Fiscal Year End", str(study_data.get("fiscal_year_end", ""))),
            ("Study Status", study_data.get("status", "").replace("_", " ").title()),
            ("Generated Date", datetime.now().strftime("%B %d, %Y")),
        ]

        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1

        # Credit Summary
        row += 1
        ws[f'A{row}'] = "Credit Summary"
        ws[f'A{row}'].font = self.SECTION_FONT
        row += 1

        credit_data = [
            ("Total QRE", calculation_result.get("total_qre", 0)),
            ("Federal Regular Credit", calculation_result.get("regular_credit", 0)),
            ("Federal ASC Credit", calculation_result.get("asc_credit", 0)),
            ("Selected Method", calculation_result.get("selected_method", "ASC").upper()),
            ("Federal Credit (Final)", calculation_result.get("federal_credit", 0)),
            ("Total State Credits", calculation_result.get("total_state_credits", 0)),
            ("Total Credits", calculation_result.get("total_credits", 0)),
        ]

        for label, value in credit_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            cell = ws[f'B{row}']
            cell.value = value
            if isinstance(value, (int, float, Decimal)) and label != "Selected Method":
                cell.number_format = self.CURRENCY_FORMAT
            if label == "Total Credits":
                cell.font = Font(bold=True, size=12, color="2E7D32")
            row += 1

        # QRE Breakdown
        row += 1
        ws[f'A{row}'] = "QRE Breakdown"
        ws[f'A{row}'].font = self.SECTION_FONT
        row += 1

        qre_data = [
            ("Wage QRE", calculation_result.get("qre_wages", 0)),
            ("Supply QRE", calculation_result.get("qre_supplies", 0)),
            ("Contract Research QRE", calculation_result.get("qre_contract", 0)),
            ("Basic Research QRE", calculation_result.get("qre_basic_research", 0)),
        ]

        for label, value in qre_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            cell = ws[f'B{row}']
            cell.value = float(value) if isinstance(value, Decimal) else value
            cell.number_format = self.CURRENCY_FORMAT
            row += 1

        # Project and Employee Summary
        row += 1
        ws[f'A{row}'] = "Study Statistics"
        ws[f'A{row}'].font = self.SECTION_FONT
        row += 1

        qualified_projects = len([p for p in projects if p.get("qualification_status") == "qualified"])
        total_employees = len(employees)

        stats_data = [
            ("Total Projects", len(projects)),
            ("Qualified Projects", qualified_projects),
            ("Total Employees", total_employees),
            ("Average Qualified Time %",
             sum(e.get("qualified_time_percentage", 0) for e in employees) / max(total_employees, 1)),
        ]

        for label, value in stats_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            cell = ws[f'B{row}']
            cell.value = value
            if "%" in label:
                cell.number_format = '0.0%'
                cell.value = value / 100 if value > 1 else value
            row += 1

        # Set column widths
        self._set_column_widths(ws, {'A': 25, 'B': 20, 'C': 15, 'D': 15, 'E': 15, 'F': 15})

    def _create_qre_summary_sheet(self, wb: Workbook, qres: List[Dict], calculation_result: Dict):
        """Create QRE summary sheet with charts."""
        ws = wb.create_sheet("QRE Summary")

        ws['A1'] = "Qualified Research Expense Summary"
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells('A1:E1')

        # QRE by Category Table
        headers = ["Category", "Item Count", "Gross Amount", "Qualified %", "QRE Amount"]
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
        self._apply_header_style(ws, row, 1, len(headers))

        # Calculate QRE by category
        categories = {
            "Wages": {"count": 0, "gross": Decimal("0"), "qualified": Decimal("0")},
            "Supplies": {"count": 0, "gross": Decimal("0"), "qualified": Decimal("0")},
            "Contract Research": {"count": 0, "gross": Decimal("0"), "qualified": Decimal("0")},
            "Basic Research": {"count": 0, "gross": Decimal("0"), "qualified": Decimal("0")},
        }

        for qre in qres:
            cat = qre.get("category", "wages").replace("_", " ").title()
            if cat not in categories:
                cat = "Wages"
            categories[cat]["count"] += 1
            categories[cat]["gross"] += Decimal(str(qre.get("gross_amount", 0)))
            categories[cat]["qualified"] += Decimal(str(qre.get("qualified_amount", 0)))

        row += 1
        total_gross = Decimal("0")
        total_qualified = Decimal("0")

        for cat_name, cat_data in categories.items():
            ws.cell(row=row, column=1, value=cat_name).border = self.THIN_BORDER
            ws.cell(row=row, column=2, value=cat_data["count"]).border = self.THIN_BORDER

            gross_cell = ws.cell(row=row, column=3, value=float(cat_data["gross"]))
            gross_cell.number_format = self.CURRENCY_FORMAT
            gross_cell.border = self.THIN_BORDER

            pct = cat_data["qualified"] / cat_data["gross"] * 100 if cat_data["gross"] else 0
            pct_cell = ws.cell(row=row, column=4, value=float(pct) / 100)
            pct_cell.number_format = self.PERCENTAGE_FORMAT
            pct_cell.border = self.THIN_BORDER

            qre_cell = ws.cell(row=row, column=5, value=float(cat_data["qualified"]))
            qre_cell.number_format = self.CURRENCY_FORMAT
            qre_cell.border = self.THIN_BORDER

            total_gross += cat_data["gross"]
            total_qualified += cat_data["qualified"]
            row += 1

        # Totals row
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = self.TOTALS_FILL
            ws.cell(row=row, column=col).font = self.TOTALS_FONT
            ws.cell(row=row, column=col).border = self.THIN_BORDER

        ws.cell(row=row, column=1, value="TOTAL")
        ws.cell(row=row, column=2, value=len(qres))
        ws.cell(row=row, column=3, value=float(total_gross)).number_format = self.CURRENCY_FORMAT
        ws.cell(row=row, column=4, value=float(total_qualified / total_gross) if total_gross else 0).number_format = self.PERCENTAGE_FORMAT
        ws.cell(row=row, column=5, value=float(total_qualified)).number_format = self.CURRENCY_FORMAT

        self._set_column_widths(ws, {'A': 20, 'B': 12, 'C': 18, 'D': 15, 'E': 18})

    def _create_employees_sheet(self, wb: Workbook, employees: List[Dict]):
        """Create employees sheet."""
        ws = wb.create_sheet("Employees")

        ws['A1'] = "Employee R&D Allocation Schedule"
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells('A1:J1')

        headers = [
            "Employee ID", "Name", "Title", "Department",
            "Total Wages", "W2 Wages", "Qualified %",
            "Qualified Wages", "Source", "CPA Reviewed"
        ]

        row = 3
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self._apply_header_style(ws, row, 1, len(headers))

        row += 1
        total_wages = Decimal("0")
        total_qualified = Decimal("0")

        for emp in employees:
            ws.cell(row=row, column=1, value=emp.get("employee_id", "")).border = self.THIN_BORDER
            ws.cell(row=row, column=2, value=emp.get("name", "")).border = self.THIN_BORDER
            ws.cell(row=row, column=3, value=emp.get("title", "")).border = self.THIN_BORDER
            ws.cell(row=row, column=4, value=emp.get("department", "")).border = self.THIN_BORDER

            tw = ws.cell(row=row, column=5, value=float(emp.get("total_wages", 0)))
            tw.number_format = self.CURRENCY_FORMAT
            tw.border = self.THIN_BORDER

            w2 = ws.cell(row=row, column=6, value=float(emp.get("w2_wages", 0)))
            w2.number_format = self.CURRENCY_FORMAT
            w2.border = self.THIN_BORDER

            pct = ws.cell(row=row, column=7, value=emp.get("qualified_time_percentage", 0) / 100)
            pct.number_format = self.PERCENTAGE_FORMAT
            pct.border = self.THIN_BORDER

            qw = ws.cell(row=row, column=8, value=float(emp.get("qualified_wages", 0)))
            qw.number_format = self.CURRENCY_FORMAT
            qw.border = self.THIN_BORDER

            ws.cell(row=row, column=9, value=emp.get("qualified_time_source", "")).border = self.THIN_BORDER
            ws.cell(row=row, column=10, value="Yes" if emp.get("cpa_reviewed") else "No").border = self.THIN_BORDER

            total_wages += Decimal(str(emp.get("w2_wages", 0)))
            total_qualified += Decimal(str(emp.get("qualified_wages", 0)))
            row += 1

        # Totals row
        for col in range(1, 11):
            ws.cell(row=row, column=col).fill = self.TOTALS_FILL
            ws.cell(row=row, column=col).font = self.TOTALS_FONT
            ws.cell(row=row, column=col).border = self.THIN_BORDER

        ws.cell(row=row, column=1, value="TOTAL")
        ws.cell(row=row, column=6, value=float(total_wages)).number_format = self.CURRENCY_FORMAT
        ws.cell(row=row, column=8, value=float(total_qualified)).number_format = self.CURRENCY_FORMAT

        self._set_column_widths(ws, {
            'A': 12, 'B': 25, 'C': 25, 'D': 15,
            'E': 15, 'F': 15, 'G': 12, 'H': 15, 'I': 15, 'J': 12
        })

    def _create_projects_sheet(self, wb: Workbook, projects: List[Dict]):
        """Create projects sheet."""
        ws = wb.create_sheet("Projects")

        ws['A1'] = "R&D Project Qualification Analysis"
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells('A1:K1')

        headers = [
            "Project Name", "Code", "Department", "Status",
            "Overall Score", "Permitted Purpose", "Tech Nature",
            "Uncertainty", "Experimentation", "Total QRE", "CPA Reviewed"
        ]

        row = 3
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self._apply_header_style(ws, row, 1, len(headers))

        row += 1
        for proj in projects:
            ws.cell(row=row, column=1, value=proj.get("name", "")).border = self.THIN_BORDER
            ws.cell(row=row, column=2, value=proj.get("code", "")).border = self.THIN_BORDER
            ws.cell(row=row, column=3, value=proj.get("department", "")).border = self.THIN_BORDER

            status = proj.get("qualification_status", "pending")
            status_cell = ws.cell(row=row, column=4, value=status.replace("_", " ").title())
            status_cell.border = self.THIN_BORDER
            if status == "qualified":
                status_cell.font = Font(color="2E7D32", bold=True)
            elif status == "not_qualified":
                status_cell.font = Font(color="C62828", bold=True)

            for col, key in enumerate([
                "overall_score", "permitted_purpose_score", "technological_nature_score",
                "uncertainty_score", "experimentation_score"
            ], 5):
                score = proj.get(key, 0)
                cell = ws.cell(row=row, column=col, value=score / 100 if score > 1 else score)
                cell.number_format = self.PERCENTAGE_FORMAT
                cell.border = self.THIN_BORDER

            qre_cell = ws.cell(row=row, column=10, value=float(proj.get("total_qre", 0)))
            qre_cell.number_format = self.CURRENCY_FORMAT
            qre_cell.border = self.THIN_BORDER

            ws.cell(row=row, column=11, value="Yes" if proj.get("cpa_reviewed") else "No").border = self.THIN_BORDER
            row += 1

        self._set_column_widths(ws, {
            'A': 35, 'B': 10, 'C': 15, 'D': 15,
            'E': 12, 'F': 15, 'G': 12, 'H': 12, 'I': 14, 'J': 15, 'K': 12
        })

    def _create_wage_qre_sheet(self, wb: Workbook, employees: List[Dict], qres: List[Dict]):
        """Create wage QRE detail sheet."""
        ws = wb.create_sheet("Wage QRE")

        ws['A1'] = "Wage QRE Schedule - IRC §41(b)(2)(A)"
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells('A1:G1')

        headers = ["Employee", "Project", "W2 Wages", "Qualified %", "Qualified Wages", "Evidence", "Notes"]
        row = 3
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self._apply_header_style(ws, row, 1, len(headers))

        row += 1
        total_qualified = Decimal("0")

        for emp in employees:
            if emp.get("qualified_wages", 0) > 0:
                ws.cell(row=row, column=1, value=emp.get("name", "")).border = self.THIN_BORDER
                ws.cell(row=row, column=2, value="Various").border = self.THIN_BORDER

                wages_cell = ws.cell(row=row, column=3, value=float(emp.get("w2_wages", 0)))
                wages_cell.number_format = self.CURRENCY_FORMAT
                wages_cell.border = self.THIN_BORDER

                pct_cell = ws.cell(row=row, column=4, value=emp.get("qualified_time_percentage", 0) / 100)
                pct_cell.number_format = self.PERCENTAGE_FORMAT
                pct_cell.border = self.THIN_BORDER

                qw_cell = ws.cell(row=row, column=5, value=float(emp.get("qualified_wages", 0)))
                qw_cell.number_format = self.CURRENCY_FORMAT
                qw_cell.border = self.THIN_BORDER

                ws.cell(row=row, column=6, value=emp.get("qualified_time_source", "")).border = self.THIN_BORDER
                ws.cell(row=row, column=7, value="").border = self.THIN_BORDER

                total_qualified += Decimal(str(emp.get("qualified_wages", 0)))
                row += 1

        # Totals row
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = self.TOTALS_FILL
            ws.cell(row=row, column=col).font = self.TOTALS_FONT
            ws.cell(row=row, column=col).border = self.THIN_BORDER

        ws.cell(row=row, column=1, value="TOTAL WAGE QRE")
        ws.cell(row=row, column=5, value=float(total_qualified)).number_format = self.CURRENCY_FORMAT

        self._set_column_widths(ws, {'A': 25, 'B': 15, 'C': 15, 'D': 12, 'E': 15, 'F': 20, 'G': 20})

    def _create_supply_qre_sheet(self, wb: Workbook, qres: List[Dict]):
        """Create supply QRE detail sheet."""
        ws = wb.create_sheet("Supply QRE")

        ws['A1'] = "Supply QRE Schedule - IRC §41(b)(2)(C)"
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells('A1:G1')

        headers = ["Description", "Vendor", "GL Account", "Gross Amount", "Qualified %", "QRE Amount", "Project"]
        row = 3
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self._apply_header_style(ws, row, 1, len(headers))

        supply_qres = [q for q in qres if q.get("category") == "supplies"]

        row += 1
        total_qre = Decimal("0")

        for qre in supply_qres:
            ws.cell(row=row, column=1, value=qre.get("supply_description", "")).border = self.THIN_BORDER
            ws.cell(row=row, column=2, value=qre.get("supply_vendor", "")).border = self.THIN_BORDER
            ws.cell(row=row, column=3, value=qre.get("gl_account", "")).border = self.THIN_BORDER

            gross_cell = ws.cell(row=row, column=4, value=float(qre.get("gross_amount", 0)))
            gross_cell.number_format = self.CURRENCY_FORMAT
            gross_cell.border = self.THIN_BORDER

            pct_cell = ws.cell(row=row, column=5, value=qre.get("qualified_percentage", 100) / 100)
            pct_cell.number_format = self.PERCENTAGE_FORMAT
            pct_cell.border = self.THIN_BORDER

            qre_cell = ws.cell(row=row, column=6, value=float(qre.get("qualified_amount", 0)))
            qre_cell.number_format = self.CURRENCY_FORMAT
            qre_cell.border = self.THIN_BORDER

            ws.cell(row=row, column=7, value=qre.get("project_name", "")).border = self.THIN_BORDER

            total_qre += Decimal(str(qre.get("qualified_amount", 0)))
            row += 1

        # Totals row
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = self.TOTALS_FILL
            ws.cell(row=row, column=col).font = self.TOTALS_FONT
            ws.cell(row=row, column=col).border = self.THIN_BORDER

        ws.cell(row=row, column=1, value="TOTAL SUPPLY QRE")
        ws.cell(row=row, column=6, value=float(total_qre)).number_format = self.CURRENCY_FORMAT

        self._set_column_widths(ws, {'A': 30, 'B': 20, 'C': 15, 'D': 15, 'E': 12, 'F': 15, 'G': 20})

    def _create_contract_qre_sheet(self, wb: Workbook, qres: List[Dict]):
        """Create contract research QRE detail sheet."""
        ws = wb.create_sheet("Contract QRE")

        ws['A1'] = "Contract Research QRE Schedule - IRC §41(b)(3)"
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells('A1:G1')

        headers = ["Contractor", "Description", "Gross Amount", "Qualified Org?", "Applicable %", "QRE Amount", "Project"]
        row = 3
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self._apply_header_style(ws, row, 1, len(headers))

        contract_qres = [q for q in qres if q.get("category") == "contract_research"]

        row += 1
        total_qre = Decimal("0")

        for qre in contract_qres:
            ws.cell(row=row, column=1, value=qre.get("contractor_name", "")).border = self.THIN_BORDER
            ws.cell(row=row, column=2, value=qre.get("description", "")).border = self.THIN_BORDER

            gross_cell = ws.cell(row=row, column=3, value=float(qre.get("gross_amount", 0)))
            gross_cell.number_format = self.CURRENCY_FORMAT
            gross_cell.border = self.THIN_BORDER

            is_qualified = qre.get("is_qualified_research_org", False)
            ws.cell(row=row, column=4, value="Yes" if is_qualified else "No").border = self.THIN_BORDER

            pct = 75 if is_qualified else 65
            pct_cell = ws.cell(row=row, column=5, value=pct / 100)
            pct_cell.number_format = self.PERCENTAGE_FORMAT
            pct_cell.border = self.THIN_BORDER

            qre_cell = ws.cell(row=row, column=6, value=float(qre.get("qualified_amount", 0)))
            qre_cell.number_format = self.CURRENCY_FORMAT
            qre_cell.border = self.THIN_BORDER

            ws.cell(row=row, column=7, value=qre.get("project_name", "")).border = self.THIN_BORDER

            total_qre += Decimal(str(qre.get("qualified_amount", 0)))
            row += 1

        # Totals row
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = self.TOTALS_FILL
            ws.cell(row=row, column=col).font = self.TOTALS_FONT
            ws.cell(row=row, column=col).border = self.THIN_BORDER

        ws.cell(row=row, column=1, value="TOTAL CONTRACT QRE")
        ws.cell(row=row, column=6, value=float(total_qre)).number_format = self.CURRENCY_FORMAT

        # Add note
        row += 2
        ws.cell(row=row, column=1, value="Note: Non-qualified organizations receive 65% credit; qualified research organizations receive 75%.")
        ws.cell(row=row, column=1).font = Font(italic=True, size=10)

        self._set_column_widths(ws, {'A': 25, 'B': 30, 'C': 15, 'D': 15, 'E': 12, 'F': 15, 'G': 20})

    def _create_federal_regular_sheet(self, wb: Workbook, calculation_result: Dict):
        """Create Federal Regular credit calculation sheet."""
        ws = wb.create_sheet("Federal Regular")

        ws['A1'] = "Federal Regular Credit Calculation - IRC §41(a)(1)"
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells('A1:D1')

        regular = calculation_result.get("federal_regular", {})

        # Current Year QRE section
        row = 3
        ws[f'A{row}'] = "Current Year QRE"
        ws[f'A{row}'].font = self.SECTION_FONT
        row += 1

        qre_items = [
            ("Wage QRE", regular.get("qre_wages", 0), "IRC §41(b)(2)(A)"),
            ("Supply QRE", regular.get("qre_supplies", 0), "IRC §41(b)(2)(C)"),
            ("Contract Research QRE", regular.get("qre_contract", 0), "IRC §41(b)(3)"),
            ("Basic Research QRE", regular.get("qre_basic_research", 0), "IRC §41(e)"),
            ("Total QRE", regular.get("total_qre", 0), ""),
        ]

        for label, value, ref in qre_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True) if label == "Total QRE" else None
            cell = ws[f'B{row}']
            cell.value = float(value) if isinstance(value, Decimal) else value
            cell.number_format = self.CURRENCY_FORMAT
            if label == "Total QRE":
                cell.font = Font(bold=True)
                cell.fill = self.TOTALS_FILL
            ws[f'C{row}'] = ref
            ws[f'C{row}'].font = Font(italic=True, size=10, color="666666")
            row += 1

        # Base Amount Calculation
        row += 1
        ws[f'A{row}'] = "Base Amount Calculation"
        ws[f'A{row}'].font = self.SECTION_FONT
        row += 1

        base_items = [
            ("Fixed-Base Percentage", regular.get("fixed_base_percentage", 0.03), "percentage"),
            ("Average Gross Receipts (4 years)", regular.get("avg_gross_receipts", 0), "currency"),
            ("Calculated Base", regular.get("calculated_base", 0), "currency"),
            ("Minimum Base (50% of QRE)", regular.get("min_base", 0), "currency"),
            ("Base Amount", regular.get("base_amount", 0), "currency"),
        ]

        for label, value, fmt in base_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True) if label == "Base Amount" else None
            cell = ws[f'B{row}']
            if fmt == "percentage":
                cell.value = float(value) if isinstance(value, Decimal) else value
                cell.number_format = self.PERCENTAGE_FORMAT
            else:
                cell.value = float(value) if isinstance(value, Decimal) else value
                cell.number_format = self.CURRENCY_FORMAT
            if label == "Base Amount":
                cell.font = Font(bold=True)
                cell.fill = self.TOTALS_FILL
            row += 1

        # Credit Calculation
        row += 1
        ws[f'A{row}'] = "Credit Calculation"
        ws[f'A{row}'].font = self.SECTION_FONT
        row += 1

        credit_items = [
            ("Excess QRE (QRE - Base)", regular.get("excess_qre", 0), "IRC §41(a)(1)"),
            ("Credit Rate", 0.20, "percentage"),
            ("Tentative Credit", regular.get("tentative_credit", 0), ""),
            ("Section 280C Reduction", regular.get("section_280c_reduction", 0), "IRC §280C(c)"),
            ("Final Regular Credit", regular.get("final_credit", 0), ""),
        ]

        for label, value, ref in credit_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True) if "Final" in label else None
            cell = ws[f'B{row}']
            if ref == "percentage" or label == "Credit Rate":
                cell.value = float(value)
                cell.number_format = self.PERCENTAGE_FORMAT
            else:
                cell.value = float(value) if isinstance(value, Decimal) else value
                cell.number_format = self.CURRENCY_FORMAT
            if "Final" in label:
                cell.font = Font(bold=True, color="2E7D32", size=12)
                cell.fill = self.TOTALS_FILL
            ws[f'C{row}'] = ref if ref != "percentage" else ""
            ws[f'C{row}'].font = Font(italic=True, size=10, color="666666")
            row += 1

        self._set_column_widths(ws, {'A': 35, 'B': 20, 'C': 15, 'D': 15})

    def _create_federal_asc_sheet(self, wb: Workbook, calculation_result: Dict):
        """Create Federal ASC calculation sheet."""
        ws = wb.create_sheet("Federal ASC")

        ws['A1'] = "Federal Alternative Simplified Credit - IRC §41(c)(4)"
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells('A1:D1')

        asc = calculation_result.get("federal_asc", {})

        # Current Year QRE
        row = 3
        ws[f'A{row}'] = "Current Year QRE"
        ws[f'A{row}'].font = self.SECTION_FONT
        row += 1

        ws[f'A{row}'] = "Total QRE"
        cell = ws[f'B{row}']
        cell.value = float(asc.get("total_qre", 0))
        cell.number_format = self.CURRENCY_FORMAT
        cell.font = Font(bold=True)
        ws[f'C{row}'] = "IRC §41(b)"
        ws[f'C{row}'].font = Font(italic=True, size=10, color="666666")
        row += 2

        # Base Amount Calculation
        ws[f'A{row}'] = "Base Amount Calculation"
        ws[f'A{row}'].font = self.SECTION_FONT
        row += 1

        base_items = [
            ("Year -1 QRE", asc.get("prior_year_1", 0)),
            ("Year -2 QRE", asc.get("prior_year_2", 0)),
            ("Year -3 QRE", asc.get("prior_year_3", 0)),
            ("Average (3 years)", asc.get("avg_prior_qre", 0)),
            ("Base Amount (50% of avg)", asc.get("base_amount", 0)),
        ]

        for label, value in base_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True) if "Base Amount" in label else None
            cell = ws[f'B{row}']
            cell.value = float(value) if isinstance(value, Decimal) else value
            cell.number_format = self.CURRENCY_FORMAT
            if "Base Amount" in label:
                cell.font = Font(bold=True)
                cell.fill = self.TOTALS_FILL
            row += 1

        # Credit Calculation
        row += 1
        ws[f'A{row}'] = "Credit Calculation"
        ws[f'A{row}'].font = self.SECTION_FONT
        row += 1

        credit_items = [
            ("Excess QRE (QRE - Base)", asc.get("excess_qre", 0), ""),
            ("Credit Rate", 0.14, "percentage"),
            ("Tentative Credit", asc.get("tentative_credit", 0), "IRC §41(c)(4)"),
            ("Section 280C Reduction", asc.get("section_280c_reduction", 0), ""),
            ("Final ASC Credit", asc.get("final_credit", 0), ""),
        ]

        for label, value, ref in credit_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True) if "Final" in label else None
            cell = ws[f'B{row}']
            if ref == "percentage" or label == "Credit Rate":
                cell.value = float(value)
                cell.number_format = self.PERCENTAGE_FORMAT
            else:
                cell.value = float(value) if isinstance(value, Decimal) else value
                cell.number_format = self.CURRENCY_FORMAT
            if "Final" in label:
                cell.font = Font(bold=True, color="2E7D32", size=12)
                cell.fill = self.TOTALS_FILL
            ws[f'C{row}'] = ref if ref != "percentage" else ""
            ws[f'C{row}'].font = Font(italic=True, size=10, color="666666")
            row += 1

        self._set_column_widths(ws, {'A': 35, 'B': 20, 'C': 15, 'D': 15})

    def _create_state_sheet(self, wb: Workbook, state_code: str, state_result: Dict):
        """Create state credit calculation sheet."""
        ws = wb.create_sheet(f"State - {state_code}")

        state_name = state_result.get("state_name", state_code)
        ws['A1'] = f"{state_name} R&D Credit Calculation"
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells('A1:C1')

        row = 3
        items = [
            ("State", state_name),
            ("Credit Type", state_result.get("credit_type", "")),
            ("State QRE", state_result.get("state_qre", 0)),
            ("Credit Rate", state_result.get("credit_rate", 0)),
            ("Base Amount", state_result.get("state_base_amount", 0)),
            ("Excess QRE", state_result.get("excess_qre", 0)),
            ("Calculated Credit", state_result.get("calculated_credit", 0)),
            ("Credit Cap", state_result.get("credit_cap", "N/A")),
            ("Final State Credit", state_result.get("final_credit", 0)),
            ("", ""),
            ("Carryforward Years", state_result.get("carryforward_years", 0)),
            ("Prior Carryforward", state_result.get("prior_carryforward", 0)),
            ("Is Refundable", "Yes" if state_result.get("is_refundable") else "No"),
        ]

        for label, value in items:
            if not label:
                row += 1
                continue
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True) if "Final" in label else None
            cell = ws[f'B{row}']

            if isinstance(value, (int, float, Decimal)) and label not in ["Carryforward Years"]:
                cell.value = float(value) if isinstance(value, Decimal) else value
                if "Rate" in label:
                    cell.number_format = self.PERCENTAGE_FORMAT
                else:
                    cell.number_format = self.CURRENCY_FORMAT
            else:
                cell.value = value

            if "Final" in label:
                cell.font = Font(bold=True, color="2E7D32", size=12)
                cell.fill = self.TOTALS_FILL
            row += 1

        self._set_column_widths(ws, {'A': 25, 'B': 20, 'C': 15})

    def _create_reconciliation_sheet(
        self,
        wb: Workbook,
        study_data: Dict,
        qres: List[Dict],
        employees: List[Dict]
    ):
        """Create reconciliation sheet."""
        ws = wb.create_sheet("Reconciliation")

        ws['A1'] = "QRE Reconciliation"
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells('A1:C1')

        total_payroll = sum(Decimal(str(e.get("total_wages", 0))) for e in employees)
        total_w2 = sum(Decimal(str(e.get("w2_wages", 0))) for e in employees)
        total_wage_qre = sum(Decimal(str(e.get("qualified_wages", 0))) for e in employees)

        row = 3
        ws[f'A{row}'] = "Payroll Reconciliation"
        ws[f'A{row}'].font = self.SECTION_FONT
        row += 1

        payroll_items = [
            ("Total Payroll (per records)", total_payroll),
            ("Total W2 Wages (employees)", total_w2),
            ("Difference", total_payroll - total_w2),
            ("", ""),
            ("Total Wage QRE", total_wage_qre),
            ("Wage QRE as % of Payroll", float(total_wage_qre / total_payroll) if total_payroll else 0),
        ]

        for label, value in payroll_items:
            if not label:
                row += 1
                continue
            ws[f'A{row}'] = label
            cell = ws[f'B{row}']
            if "%" in label:
                cell.value = value
                cell.number_format = self.PERCENTAGE_FORMAT
            else:
                cell.value = float(value) if isinstance(value, Decimal) else value
                cell.number_format = self.CURRENCY_FORMAT
            row += 1

        row += 1
        ws[f'A{row}'] = "QRE Category Reconciliation"
        ws[f'A{row}'].font = self.SECTION_FONT
        row += 1

        supply_qre = sum(Decimal(str(q.get("qualified_amount", 0))) for q in qres if q.get("category") == "supplies")
        contract_qre = sum(Decimal(str(q.get("qualified_amount", 0))) for q in qres if q.get("category") == "contract_research")
        total_qre = total_wage_qre + supply_qre + contract_qre

        qre_items = [
            ("Total Wage QRE", total_wage_qre),
            ("Total Supply QRE", supply_qre),
            ("Total Contract QRE", contract_qre),
            ("Total QRE", total_qre),
        ]

        for label, value in qre_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True) if "Total QRE" == label else None
            cell = ws[f'B{row}']
            cell.value = float(value) if isinstance(value, Decimal) else value
            cell.number_format = self.CURRENCY_FORMAT
            if "Total QRE" == label:
                cell.font = Font(bold=True)
                cell.fill = self.TOTALS_FILL
            row += 1

        self._set_column_widths(ws, {'A': 35, 'B': 20, 'C': 15})

    def _create_sanity_checks_sheet(
        self,
        wb: Workbook,
        study_data: Dict,
        calculation_result: Dict
    ):
        """Create sanity checks sheet."""
        ws = wb.create_sheet("Sanity Checks")

        ws['A1'] = "R&D Credit Sanity Checks"
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells('A1:D1')

        total_qre = float(calculation_result.get("total_qre", 0))
        federal_credit = float(calculation_result.get("federal_credit", 0))
        qre_wages = float(calculation_result.get("qre_wages", 0))
        qre_supplies = float(calculation_result.get("qre_supplies", 0))

        credit_pct = federal_credit / total_qre if total_qre else 0
        wage_pct = qre_wages / total_qre if total_qre else 0
        supply_pct = qre_supplies / total_qre if total_qre else 0

        checks = [
            ("Credit as % of QRE", credit_pct, "8% - 14%", 0.08 <= credit_pct <= 0.14),
            ("Wage QRE as % of Total", wage_pct, "60% - 80%", 0.60 <= wage_pct <= 0.80),
            ("Supply QRE as % of Total", supply_pct, "5% - 20%", supply_pct <= 0.20),
            ("Total QRE > $0", total_qre, "> $0", total_qre > 0),
            ("Federal Credit > $0", federal_credit, "> $0", federal_credit > 0),
        ]

        headers = ["Check", "Value", "Expected Range", "Status"]
        row = 3
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self._apply_header_style(ws, row, 1, len(headers))

        row += 1
        for check_name, value, expected, passed in checks:
            ws.cell(row=row, column=1, value=check_name).border = self.THIN_BORDER

            value_cell = ws.cell(row=row, column=2)
            if isinstance(value, float) and value < 1:
                value_cell.value = value
                value_cell.number_format = self.PERCENTAGE_FORMAT
            else:
                value_cell.value = value
                value_cell.number_format = self.CURRENCY_FORMAT
            value_cell.border = self.THIN_BORDER

            ws.cell(row=row, column=3, value=expected).border = self.THIN_BORDER

            status_cell = ws.cell(row=row, column=4, value="PASS" if passed else "REVIEW")
            status_cell.border = self.THIN_BORDER
            if passed:
                status_cell.font = Font(bold=True, color="2E7D32")
                status_cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            else:
                status_cell.font = Font(bold=True, color="E65100")
                status_cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

            row += 1

        self._set_column_widths(ws, {'A': 30, 'B': 15, 'C': 15, 'D': 12})

    def _create_form_6765_sheet(self, wb: Workbook, study_data: Dict, calculation_result: Dict):
        """Create Form 6765 data sheet."""
        ws = wb.create_sheet("Form 6765 Data")

        ws['A1'] = "IRS Form 6765 - Credit for Increasing Research Activities"
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells('A1:D1')

        row = 3
        ws[f'A{row}'] = "Entity Information"
        ws[f'A{row}'].font = self.SECTION_FONT
        row += 1

        entity_info = [
            ("Name", study_data.get("entity_name", "")),
            ("EIN", study_data.get("ein", "")),
            ("Tax Year", study_data.get("tax_year", "")),
        ]

        for label, value in entity_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1

        row += 1
        ws[f'A{row}'] = "Section A - Regular Credit (Key Lines)"
        ws[f'A{row}'].font = self.SECTION_FONT
        row += 1

        regular = calculation_result.get("federal_regular", {})
        section_a = [
            ("Line 5 - Wages for qualified services", regular.get("qre_wages", 0)),
            ("Line 6 - Cost of supplies", regular.get("qre_supplies", 0)),
            ("Line 8 - Contract research expenses", regular.get("qre_contract", 0)),
            ("Line 9 - Total qualified research expenses", regular.get("total_qre", 0)),
            ("Line 10 - Fixed-base percentage", regular.get("fixed_base_percentage", 0)),
            ("Line 13 - Excess QRE over base", regular.get("excess_qre", 0)),
            ("Line 14 - Credit (Line 13 × 20%)", regular.get("tentative_credit", 0)),
        ]

        for label, value in section_a:
            ws[f'A{row}'] = label
            cell = ws[f'B{row}']
            if "percentage" in label.lower():
                cell.value = float(value) if isinstance(value, Decimal) else value
                cell.number_format = self.PERCENTAGE_FORMAT
            else:
                cell.value = float(value) if isinstance(value, Decimal) else value
                cell.number_format = self.CURRENCY_FORMAT
            row += 1

        row += 1
        ws[f'A{row}'] = "Section B - Alternative Simplified Credit (Key Lines)"
        ws[f'A{row}'].font = self.SECTION_FONT
        row += 1

        asc = calculation_result.get("federal_asc", {})
        section_b = [
            ("Line 35 - Total qualified research expenses", asc.get("total_qre", 0)),
            ("Line 37 - Average QRE (3 prior years) ÷ 3", asc.get("avg_prior_qre", 0)),
            ("Line 38 - Base amount (Line 37 × 50%)", asc.get("base_amount", 0)),
            ("Line 39 - Excess QRE over base", asc.get("excess_qre", 0)),
            ("Line 40 - Credit (Line 39 × 14%)", asc.get("tentative_credit", 0)),
        ]

        for label, value in section_b:
            ws[f'A{row}'] = label
            cell = ws[f'B{row}']
            cell.value = float(value) if isinstance(value, Decimal) else value
            cell.number_format = self.CURRENCY_FORMAT
            row += 1

        row += 1
        ws[f'A{row}'] = "Section C - Current Year Credit"
        ws[f'A{row}'].font = self.SECTION_FONT
        row += 1

        selected_method = calculation_result.get("selected_method", "asc")
        section_c = [
            ("Line 42 - Credit from Section A or B", calculation_result.get("federal_credit", 0) / 0.79 if calculation_result.get("federal_credit", 0) else 0),
            ("Line 44 - Section 280C reduction (21%)", calculation_result.get("section_280c_reduction", 0)),
            ("Line 45 - Final credit", calculation_result.get("federal_credit", 0)),
        ]

        for label, value in section_c:
            ws[f'A{row}'] = label
            cell = ws[f'B{row}']
            cell.value = float(value) if isinstance(value, Decimal) else value
            cell.number_format = self.CURRENCY_FORMAT
            if "Final" in label:
                cell.font = Font(bold=True, color="2E7D32", size=12)
                cell.fill = self.TOTALS_FILL
            row += 1

        row += 1
        ws[f'A{row}'] = f"Selected Method: {selected_method.upper()}"
        ws[f'A{row}'].font = Font(bold=True, color="1F4E79")

        self._set_column_widths(ws, {'A': 50, 'B': 20, 'C': 15, 'D': 15})

    def _create_raw_imports_sheet(self, wb: Workbook, raw_data: Dict):
        """Create raw imports sheet (read-only reference)."""
        ws = wb.create_sheet("Raw Imports")

        ws['A1'] = "Raw Imported Data - DO NOT MODIFY"
        ws['A1'].font = Font(bold=True, size=14, color="C62828")
        ws.merge_cells('A1:E1')

        row = 3

        # GL Data
        if raw_data.get("gl_data"):
            ws[f'A{row}'] = "General Ledger Import"
            ws[f'A{row}'].font = self.SECTION_FONT
            row += 1

            headers = ["Account", "Description", "Debit", "Credit", "Source"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header)
            self._apply_header_style(ws, row, 1, len(headers))
            row += 1

            for gl_row in raw_data["gl_data"][:100]:  # Limit to 100 rows
                for col, value in enumerate(gl_row, 1):
                    ws.cell(row=row, column=col, value=value).border = self.THIN_BORDER
                row += 1
            row += 1

        # Payroll Data
        if raw_data.get("payroll_data"):
            ws[f'A{row}'] = "Payroll Import"
            ws[f'A{row}'].font = self.SECTION_FONT
            row += 1

            headers = ["Employee", "Wages", "Bonus", "Total", "Source"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header)
            self._apply_header_style(ws, row, 1, len(headers))
            row += 1

            for payroll_row in raw_data["payroll_data"][:100]:
                for col, value in enumerate(payroll_row, 1):
                    ws.cell(row=row, column=col, value=value).border = self.THIN_BORDER
                row += 1

        # Protect sheet
        ws.protection.sheet = True
        ws.protection.password = 'readonly'

        self._set_column_widths(ws, {'A': 20, 'B': 30, 'C': 15, 'D': 15, 'E': 15})
